"""
Polarization Curve Fitter — Publication-Grade Streamlit App
(Robust local Tafel overlays + Tafel intersection i_corr + adjustable detection)
===============================================================================
- Vectorized sliding-window regressions with curvature & diffusion guards
- Huber refinement (IRLS) and Theil–Sen on the chosen Tafel windows
- Local Tafel dashed lines drawn as straight segments (with optional faint extension)
- i_corr annotated from Tafel intersection of local anodic/cathodic lines (toggleable)
- Lean global optimization and adaptive plotting
- Stable widget keys; force-model selection via session-backed key

FIXES (v2):
  - Cathodic proximity score reversed: now rewards windows 80–250 mV from E_corr
    (true Tafel region), not windows near E_corr (mixed-control zone).
  - Cathodic scoring: R²² weighting added; curvature penalty incorporated directly.
  - Cathodic i_corr: uses max(tafel_extrap, near_Ecorr) rather than the
    incorrectly capped min() which under-estimated i_corr.
  - Anodic proximity tau widened from 0.12 V → 0.20 V to accept active regions
    spanning 100–200 mV (e.g. stainless, Ni alloys).
  - Anodic fallback ok mask retains beta_ok to prevent unphysical slopes.
  - Linear-scale panel (Panel C) ylim now zooms to ±2× the local i range within
    ±150 mV of E_corr, avoiding passive/transpassive peaks crushing the view.
  - diff_ok guard relaxed: threshold lowered from 0.35→0.20 slope_mag to avoid
    rejecting valid cathodic windows with mild slope variation.

Run:
    streamlit run app.py
"""

import streamlit as st
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
from matplotlib.ticker import AutoMinorLocator
import io, zipfile, warnings, traceback, re
from itertools import groupby
from scipy.optimize import differential_evolution, minimize
from scipy.signal import savgol_filter, find_peaks
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors as rl_colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, Image as RLImage, HRFlowable, KeepTogether)

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Polarization Curve Fitter", page_icon="⚡",
                   layout="wide", initial_sidebar_state="expanded")
st.markdown("""
<style>
  .main-header{font-size:2rem;font-weight:700;color:#1a3a5c;
    border-bottom:3px solid #2e86de;padding-bottom:8px;margin-bottom:1rem}
  div[data-testid="metric-container"]{
    background:#f0f4ff;border-left:3px solid #2e86de;
    border-radius:6px;padding:8px 12px}
</style>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS & CONFIG
# ─────────────────────────────────────────────────────────────────────────────
TINY    = 1e-30
PALETTE = ["#2e86de","#e84393","#27ae60","#e67e22","#8e44ad","#16a085","#c0392b"]

REGION_COLORS = {
    "cathodic":     ("#6baed6", 0.14),
    "active":       ("#fd8d3c", 0.22),
    "passive":      ("#74c476", 0.16),
    "transpassive": ("#9e9ac8", 0.18),
}

MATERIALS = {
    "Carbon Steel / Iron":   (27.92, 7.87),
    "304 Stainless Steel":   (25.10, 7.90),
    "316 Stainless Steel":   (25.56, 8.00),
    "Copper":                (31.77, 8.96),
    "Aluminum":              ( 8.99, 2.70),
    "Nickel":                (29.36, 8.91),
    "Titanium":              (11.99, 4.51),
    "Zinc":                  (32.69, 7.14),
}

# Global detection settings (overridden by sidebar "Advanced")
CFG = dict(
    cath_guard=0.020,      # V — min distance of cathodic window from Ecorr
    anod_guard=0.010,      # V — min distance of anodic window from Ecorr
    curvature_max=40.0,    # max |d2 log|i| / dE^2| inside window
    lin_frac=0.70,         # fraction of derivative signs consistent with slope
    min_w_cat=4,           # minimum window length (cathodic)
    min_w_ano=3,           # minimum window length (anodic)
    # Tafel slope range (V/dec). Physical upper limits based on BV theory:
    # ba = RT/(alpha*F). alpha<0.09 -> ba>280mV/dec is unphysical.
    # Old beta_max_a=0.400 let passive-region slopes (ba~300-400mV) pass
    # the beta_ok gate, producing the ba=384mV/dec artefact seen in image.
    beta_min=0.020,        # 20 mV/dec lower physical limit
    beta_max_c=0.280,      # 280 mV/dec cathodic cap (was 350)
    beta_max_a=0.250,      # 250 mV/dec anodic cap   (was 400 — this was the bug)
)

# ─────────────────────────────────────────────────────────────────────────────
# HELPER MATH
# ─────────────────────────────────────────────────────────────────────────────
def slog(x): return np.log10(np.maximum(np.abs(x), TINY))

def sig(x, k=40.0):
    xk = np.clip(k * x, -60, 60)
    return np.where(xk >= 0, 1.0 / (1.0 + np.exp(-xk)),
                    np.exp(xk) / (1.0 + np.exp(xk)))

def sm(y, w=11, p=3):
    n = len(y)
    w = min(w, n if n % 2 == 1 else n - 1)
    w = max(5, w) if w >= 5 else n
    if w > n or w < 5: return y.copy()
    return savgol_filter(y, w, min(p, w - 1), mode="interp")

def r2_score(yt, yp):
    sr = np.sum((yt - yp) ** 2)
    st = np.sum((yt - np.mean(yt)) ** 2)
    return float(max(0.0, 1.0 - sr / st)) if st > 1e-30 else 0.0

def aicc(n, k, sse):
    if n <= k + 1 or sse <= 0: return 1e30
    return n * np.log(sse / n) + 2 * k + (2 * k * (k + 1)) / max(n - k - 1, 1)

def downsample_uniform(x, y, max_pts=400):
    if len(x) <= max_pts: return x, y
    idx = np.linspace(0, len(x)-1, max_pts).astype(int)
    return x[idx], y[idx]

# ─────────────────────────────────────────────────────────────────────────────
# FAST SLIDING REGRESSION (VECTORIZED)
# ─────────────────────────────────────────────────────────────────────────────
def _sliding_regress_full(x, y, min_len=4, max_len=25):
    n = len(x)
    if n < min_len:
        return (np.array([], int), np.array([], int),
                np.array([]), np.array([]), np.array([]))
    Sx  = np.cumsum(x); Sy  = np.cumsum(y)
    Sxx = np.cumsum(x*x); Sxy = np.cumsum(x*y); Syy = np.cumsum(y*y)

    starts_all, ends_all, slopes_all, inters_all, r2_all = [], [], [], [], []
    wmax = min(max_len, n)
    for w in range(min_len, wmax+1):
        i0 = np.arange(0, n - w + 1)
        i1 = i0 + w - 1
        def segsum(csum): return csum[i1] - np.concatenate(([0.0], csum[i0[:-1]]))
        sum_x, sum_y = segsum(Sx), segsum(Sy)
        sum_xx, sum_xy, sum_yy = segsum(Sxx), segsum(Sxy), segsum(Syy)

        w_f = float(w)
        mx, my = sum_x / w_f, sum_y / w_f
        denom = sum_xx - w_f * mx * mx
        slope = np.where(np.abs(denom) > 1e-18, (sum_xy - w_f * mx * my) / denom, 0.0)
        intercept = my - slope * mx

        SSE = (sum_yy - 2.0*intercept*sum_y - 2.0*slope*sum_xy
               + (intercept**2)*w_f + 2.0*intercept*slope*sum_x + (slope**2)*sum_xx)
        SST = sum_yy - w_f * my * my
        R2 = np.where(SST > 1e-18, 1.0 - SSE / SST, 0.0)
        R2 = np.clip(R2, 0.0, 1.0)

        starts_all.append(i0); ends_all.append(i1 + 1)
        slopes_all.append(slope); inters_all.append(intercept); r2_all.append(R2)

    return (np.concatenate(starts_all), np.concatenate(ends_all),
            np.concatenate(slopes_all), np.concatenate(inters_all),
            np.concatenate(r2_all))

# ─────────────────────────────────────────────────────────────────────────────
# CURVE TYPE REGISTRY
# ─────────────────────────────────────────────────────────────────────────────
PARAM_NAMES = ["Ecorr","icorr","ba","bc","Epass","k_pass","ip","Etrans","k_trans","itrans","iL"]
NP = 11

class CT:
    A  = "A"; AD = "AD"; P  = "P"; PT = "PT"; F  = "F"
    INFO = {
        "A":  ("Active",                    [0,1,2,3],             4),
        "AD": ("Active + Diffusion",        [0,1,2,3,10],          5),
        "P":  ("Active–Passive",            [0,1,2,3,4,5,6],       7),
        "PT": ("Active–Passive–Transpassive",[0,1,2,3,4,5,6,7,8,9],10),
        "F":  ("Full (PT+Diffusion)",       list(range(NP)),       11),
    }
    ALL    = ["A","AD","P","PT","F"]
    SIMPLE = ["A","AD"]
    PASS   = ["P","PT","F"]
    TRANS  = ["PT","F"]

    @staticmethod
    def idx(ct):   return CT.INFO.get(ct, CT.INFO["A"])[1]
    @staticmethod
    def nfree(ct): return CT.INFO.get(ct, CT.INFO["A"])[2]
    @staticmethod
    def name(ct):  return CT.INFO.get(ct, ("?", [], 0))[0]

# ─────────────────────────────────────────────────────────────────────────────
# PHYSICS MODEL
# ─────────────────────────────────────────────────────────────────────────────
def pol_model(E, p, ct="PT"):
    E    = np.asarray(E, float)
    Ec   = p[0]; ic = p[1]; ba = max(p[2], 1e-6); bc = max(p[3], 1e-6)
    Ep   = p[4]; kp = max(p[5], 0.001); ip = p[6]
    Et   = p[7]; kt = max(p[8], 0.001); it = p[9]
    iL   = max(p[10], 1e-30)
    eta  = E - Ec
    ik_cat = ic * np.exp(np.clip(-2.303 * eta / bc, -60, 60))
    i_cat = ik_cat / (1.0 + ik_cat / iL) if ct in ("AD","F") else ik_cat
    i_act = ic * np.exp(np.clip(2.303 * eta / ba, -60, 60))
    if ct in CT.SIMPLE: return i_act - i_cat
    w_p   = sig(E - Ep, 1.0 / kp)
    i_ano = (1.0 - w_p) * i_act + w_p * ip
    if ct == "P": return i_ano - i_cat
    w_t   = sig(E - Et, 1.0 / kt)
    i_tp  = ip + it * np.exp(np.clip(2.303 * (E - Et) / ba, -60, 60))
    i_ano = (1.0 - w_t) * i_ano + w_t * i_tp
    return i_ano - i_cat

# ─────────────────────────────────────────────────────────────────────────────
# STAGE 1 — E_corr DETECTION
# ─────────────────────────────────────────────────────────────────────────────
def detect_ecorr(E, i):
    E = np.asarray(E, float); i = np.asarray(i, float)
    si = np.argsort(E); Es = E[si]; is_ = i[si]
    sc = np.where(np.diff(np.sign(is_)))[0]
    if len(sc) == 0:
        idx = int(np.argmin(np.abs(is_))); return float(Es[idx]), int(si[idx])
    crossings = []
    for k in sc:
        denom = is_[k+1] - is_[k]
        if abs(denom) < TINY: continue
        Ec = float(Es[k] - is_[k] * (Es[k+1] - Es[k]) / denom)
        goes_anodic = (is_[k] < 0 and is_[k+1] > 0)
        crossings.append((Ec, int(si[k]), goes_anodic))
    anodic = [(Ec, idx) for Ec, idx, ga in crossings if ga]
    if anodic: return min(anodic, key=lambda x: x[0])
    best = min(crossings, key=lambda x: x[0]); return best[0], best[1]

# ─────────────────────────────────────────────────────────────────────────────
# ROBUST LINE REFINEMENT
# ─────────────────────────────────────────────────────────────────────────────
def _huber_fit(x, y, slope, intercept, iters=3, c=1.345):
    for _ in range(iters):
        r = y - (slope * x + intercept)
        s = np.median(np.abs(r)) * 1.4826 + 1e-12
        w = np.ones_like(r)
        t = np.abs(r) / (c * s + 1e-12)
        w[t > 1] = (c * s) / (np.abs(r[t > 1]) + 1e-12)
        X = np.vstack([np.ones_like(x), x]).T
        W = np.diag(w)
        try:
            b, a = np.linalg.lstsq(W @ X, W @ y, rcond=None)[0]
            intercept, slope = float(b), float(a)
        except Exception:
            break
    return slope, intercept

def _theil_sen(x, y):
    m = len(x)
    if m < 2:
        return 0.0, float(np.median(y)) if m else (0.0, 0.0)
    i_idx, j_idx = np.triu_indices(m, k=1)
    dx = x[j_idx] - x[i_idx]
    valid = np.abs(dx) > 1e-15
    slopes = (y[j_idx][valid] - y[i_idx][valid]) / dx[valid]
    slope = float(np.median(slopes)) if len(slopes) else 0.0
    intercept = float(np.median(y - slope * x))
    return slope, intercept

# ─────────────────────────────────────────────────────────────────────────────
# STAGE 2 — CATHODIC BRANCH FIT (ROBUST LOCAL LINE)
# FIX: Corrected proximity scoring, curvature weighting, i_corr estimation
# ─────────────────────────────────────────────────────────────────────────────
def fit_cathodic(E, i, Ecorr):
    """
    Find the cathodic Tafel window using hard beta enforcement.

    Key design decisions:
    - beta_ok is a HARD GATE: windows outside the physical Tafel slope range
      are completely excluded (score=0), not just penalised. If all windows
      fail, we fall back to the steepest valid-direction window.
    - Proximity scoring uses a Gaussian centred 150 mV below Ecorr, rewarding
      the true Tafel zone and penalising both mixed-control (too close) and
      diffusion-plateau (too flat/far) regions.
    - Points within the top 1.5 decades of log|i| closest to Ecorr are trimmed
      before window search to avoid mixed-control corruption.
    """
    cat = i < 0
    if np.sum(cat) < 4:
        return dict(bc=0.120, icorr=1e-8, iL=1e-2, has_diff=False, r2=0.0)

    Ec  = E[cat]; lgi = slog(i[cat])
    si  = np.argsort(Ec); Ec, lgi = Ec[si], lgi[si]

    # Strict guard: exclude points within cath_guard of Ecorr
    base_mask = Ec < (Ecorr - CFG["cath_guard"])
    if np.sum(base_mask) < CFG["min_w_cat"]:
        base_mask = np.ones_like(Ec, bool)

    Ex, Yx = Ec[base_mask], lgi[base_mask]
    if len(Ex) < CFG["min_w_cat"]:
        return dict(bc=0.120, icorr=max(10**np.min(Yx), 1e-12), iL=1e-2,
                    has_diff=False, r2=0.0, E_cat=Ec, lgi_cat=lgi)

    # Remove the diffusion plateau top (keeps Tafel + transition, drops flat plateau)
    lgi_max_cat = np.max(Yx)
    trim_mask = Yx < (lgi_max_cat - 0.3)
    if np.sum(trim_mask) >= CFG["min_w_cat"]:
        Ex, Yx = Ex[trim_mask], Yx[trim_mask]

    # Smooth & derivatives (use full trimmed range for smoothing)
    w_sm = max(5, min(11, len(Ex) // 2 * 2 - 1))
    Y_sm = savgol_filter(Yx, w_sm, 3, mode="interp")
    dY   = np.gradient(Y_sm, Ex)
    d2Y  = np.gradient(dY, Ex)

    # ── Steep-region pre-filter ───────────────────────────────────────────────
    # For diffusion-limited cathodic branches the data contains:
    #   [transition zone: |dY/dE| is moderate] → [Tafel zone: |dY/dE| is maximum]
    # The sliding window over the full range finds the smooth transition (low
    # curvature, moderate slope) rather than the steep Tafel zone (high curvature
    # because it borders the mixed-control zone near Ecorr).
    # Fix: restrict the window search to points where the local slope magnitude
    # exceeds 50% of its maximum value — this selects only the Tafel zone.
    abs_dY    = np.abs(dY)
    max_abs_dY = float(np.max(abs_dY)) if len(abs_dY) > 0 else 1.0
    steep_pre  = abs_dY > 0.50 * max_abs_dY
    if np.sum(steep_pre) >= CFG["min_w_cat"]:
        Ex_sw, Yx_sw = Ex[steep_pre], Yx[steep_pre]
        dY_sw  = dY[steep_pre]
        d2Y_sw = d2Y[steep_pre]
    else:
        # Fallback: use full trimmed range if steep region is too small
        Ex_sw, Yx_sw, dY_sw, d2Y_sw = Ex, Yx, dY, d2Y

    s_idx, e_idx, slope, intercept, R2 = _sliding_regress_full(
        Ex_sw, Yx_sw, min_len=CFG["min_w_cat"], max_len=25)
    if len(slope) == 0:
        return dict(bc=0.120, icorr=1e-9, iL=1e-2, has_diff=False, r2=0.0,
                    E_cat=Ec, lgi_cat=lgi)

    # Per-window quality metrics on the (possibly pre-filtered) data
    max_d2 = np.array([np.max(np.abs(d2Y_sw[s:e])) for s, e in zip(s_idx, e_idx)])

    def _mono_frac(s, e):
        seg = dY_sw[s:e]
        return float(np.mean(seg < 0)) if len(seg) else 0.0
    mono = np.array([_mono_frac(s, e) for s, e in zip(s_idx, e_idx)])

    invm     = np.where(np.abs(slope) > 1e-12, 1.0 / np.abs(slope), np.inf)
    # HARD GATE: beta_ok is a strict physical constraint, not a scoring term
    beta_ok  = (slope < 0) & (invm > CFG["beta_min"]) & (invm < CFG["beta_max_c"])
    mono_ok  = mono >= CFG["lin_frac"]
    # Note: curvature is NOT used as a hard gate — for diffusion-limited cathodic
    # branches the second derivative is large everywhere (sigmoid morphology), so
    # curvature_max would eliminate ALL windows.  Instead, curvature feeds only
    # into the scoring penalty below.

    # Primary candidates: beta_ok + monotone + decent R²
    ok = beta_ok & mono_ok & (R2 > 0.85)
    if not np.any(ok):
        ok = beta_ok & (R2 > 0.75)
    if not np.any(ok):
        ok = beta_ok.copy()
    if not np.any(ok):
        ok = slope < 0
        if not np.any(ok):
            ok = np.ones(len(slope), bool)

    # ── Scoring ───────────────────────────────────────────────────────────────
    # For diffusion-limited cathodic branches the morphology is:
    #   [diffusion plateau] → [curved transition] → [steep Tafel] → [mixed control]
    # The plateau and transition have SHALLOW slopes → eliminated by beta_ok.
    # Within beta_ok windows, the TRUE Tafel region is the STEEPEST one.
    # We therefore weight by (steep_bonus)² to sharply prefer the Tafel region.
    # Curvature enters only as a soft penalty (exp term), never as a hard gate.
    curv_pen   = np.exp(-max_d2 / max(CFG["curvature_max"], 10.0))
    win_len    = (e_idx - s_idx).astype(float)
    len_bonus  = np.log1p(win_len) / np.log1p(25.0)

    abs_slope       = np.abs(slope)
    max_beta_slope  = float(np.max(abs_slope[beta_ok])) if np.any(beta_ok) else 1.0
    steep_bonus     = abs_slope / max(max_beta_slope, 1e-9)   # [0,1]

    # steep_bonus² dominates: steepest (= Tafel) window wins decisively
    score = (R2 ** 2) * np.clip(mono, 0.2, 1.0) * curv_pen \
            * (0.3 + 0.1 * len_bonus + 0.6 * steep_bonus ** 2)

    score[~beta_ok] = 0.0                          # hard zero: wrong direction/beta
    score[beta_ok & ~mono_ok] *= 0.40              # soft penalty: non-monotone windows

    if np.all(score == 0.0) or not np.any(ok):
        neg_slope_mag = np.where(slope < 0, np.abs(slope), 0.0)
        k_best = int(np.argmax(neg_slope_mag))
    else:
        k_best = int(np.argmax(score))

    s0, s1   = int(s_idx[k_best]), int(e_idx[k_best])
    sl_ols, b_ols = float(slope[k_best]), float(intercept[k_best])

    # Theil–Sen + Huber; pick best R²
    sl_ts, b_ts = _theil_sen(Ex[s0:s1], Yx[s0:s1])
    sl_hb, b_hb = _huber_fit(Ex[s0:s1], Yx[s0:s1], sl_ols, b_ols)

    def _r2_line(sl, b):
        return r2_score(Yx[s0:s1], sl * Ex[s0:s1] + b)
    cand = [(sl_ts, b_ts, _r2_line(sl_ts, b_ts)),
            (sl_hb, b_hb, _r2_line(sl_hb, b_hb)),
            (sl_ols, b_ols, _r2_line(sl_ols, b_ols))]
    sl_ref, b_ref, r2_win = max(cand, key=lambda t: t[2])

    # Ensure the refined slope is physically sensible
    if sl_ref >= 0 or abs(1.0 / sl_ref) > CFG["beta_max_c"] * 1.5:
        sl_ref, b_ref = sl_ols, b_ols  # revert to OLS if refinement drifted

    bc = min(abs(1.0 / sl_ref), CFG["beta_max_c"]) if abs(sl_ref) > 1e-9 else 0.120

    # Tafel extrapolation to Ecorr
    icorr_tafel = float(10 ** (b_ref + sl_ref * Ecorr)) if abs(sl_ref) > 1e-9 else 1e-12

    # Cross-check near Ecorr; use Tafel extrap as primary (it is the physical value)
    near = np.abs(Ec - Ecorr) < 0.050
    if np.any(near):
        icorr_near = float(np.percentile(np.abs(i[cat][si][near]), 25))
        icorr = max(icorr_tafel, icorr_near * 0.5)
    else:
        icorr = icorr_tafel
    icorr = max(icorr, 1e-15)

    # Diffusion limit detection
    iL, has_diff = None, False
    if len(Ec) > 8:
        lgi_sm = savgol_filter(lgi, max(5, min(9, len(Ec) // 2 * 2 - 1)), 3, mode="interp")
        dlg    = np.abs(np.gradient(lgi_sm, Ec))
        flat   = dlg < max(np.percentile(dlg, 30), 0.5)
        runs   = [(k, list(g)) for k, g in groupby(enumerate(flat), key=lambda x: x[1]) if k]
        if runs:
            idxs = [s[0] for s in max(runs, key=lambda x: len(x[1]))[1]]
            if len(idxs) >= 3 and abs(Ec[idxs[-1]] - Ec[idxs[0]]) > 0.03:
                iL = float(np.median(np.abs(i[cat][si][idxs]))); has_diff = True
    if iL is None:
        iL = icorr * 1e4

    return dict(
        bc=bc, icorr=icorr, iL=iL, has_diff=has_diff,
        r2=float(r2_win),
        E_cat=Ec, lgi_cat=lgi,
        slope_c=sl_ref, intercept_c=b_ref, win_c=(Ex[s0], Ex[s1 - 1])
    )

# ─────────────────────────────────────────────────────────────────────────────
# STAGE 3 — ANODIC BRANCH FIT (ROBUST LOCAL LINE)
# FIX: Widened proximity tau, corrected fallback ok mask
# ─────────────────────────────────────────────────────────────────────────────
def fit_anodic(E, i, Ecorr):
    """
    Find the anodic Tafel (active dissolution) window using hard beta enforcement.

    Key design decisions:
    - beta_ok is a HARD GATE: flat passive/plateau windows (slope~0, 1/slope>>400mV)
      get score=0 regardless of their R². This was the main failure mode where the
      algorithm mistakenly selected the active dissolution peak or passive plateau
      (high R², nearly flat) over the true steep Tafel region.
    - Passive detection is run FIRST to bound the Tafel search to E < E_pass.
    - If all windows fail beta_ok, pick the steepest positive-slope window
      (most Tafel-like) rather than the flattest high-R² one.
    - Proximity tau widened to 0.20 V to accept active regions spanning 100–200 mV.
    """
    ano = i > 0
    if np.sum(ano) < 4:
        return dict(ba=0.060, has_passive=False, Epass=None, ip=1e-6,
                    has_trans=False, Etrans=None, r2=0.0)

    Ea  = E[ano]; lgia = slog(i[ano])
    si  = np.argsort(Ea); Ea, lgia = Ea[si], lgia[si]

    # ── Passive / Transpassive detection (run BEFORE Tafel search so we can
    #    bound the active region search range to E < E_pass) ─────────────────
    has_passive = False; Epass = None; ip = 1e-6
    has_trans   = False; Etrans = None
    Epeak_detected = None

    if len(Ea) > 6:
        # Smooth first for robust derivative and peak detection
        # Cap at 9: a 15-point window over-smooths the narrow active peak
        # (e.g. 8 pts wide), shifting the detected Epeak rightward into the
        # passive transition zone and inflating ba (image gave ba=184 vs ~95 true).
        w_pk = max(5, min(9, len(Ea) // 2 * 2 - 1))
        lgia_pk = savgol_filter(lgia, w_pk, 3, mode="interp")
        dY_pk   = np.gradient(lgia_pk, Ea)

        # PRIMARY: derivative sign change (+→-) above Ecorr = active dissolution peak.
        # This catches even very shallow peaks (< 0.1 decade) that find_peaks misses.
        sign_chg = np.where(np.diff(np.sign(dY_pk)) < 0)[0]   # rising→falling
        cands = [Ea[k] for k in sign_chg
                 if Ea[k] > Ecorr + 0.005 and Ea[k] < Ea[-1] - 0.05]
        if cands:
            # Prefer the first candidate within the expected active zone
            # (within 200 mV of Ecorr). This prevents noise sign-changes in
            # the passive/transpassive region from being mistaken for the
            # active dissolution peak (which caused ba=184 on the image data).
            active_cands = [c for c in cands if c < Ecorr + 0.200]
            Epeak_detected = float(min(active_cands)) if active_cands else float(min(cands))

        # SECONDARY: find_peaks with reduced prominence
        # Same 200mV active-zone filter applied for consistency.
        if Epeak_detected is None:
            pks, _ = find_peaks(lgia_pk, prominence=0.08, distance=2)
            pks_active = [p for p in pks
                          if Ecorr + 0.005 < Ea[p] < Ecorr + 0.200]
            if pks_active:
                pk = min(pks_active, key=lambda p: abs(Ea[p] - Ecorr))
                Epeak_detected = float(Ea[pk])
            elif len(pks) > 0:
                pks_above = [p for p in pks if Ea[p] > Ecorr + 0.005]
                if pks_above:
                    pk = min(pks_above, key=lambda p: abs(Ea[p] - Ecorr))
                    Epeak_detected = float(Ea[pk])

    if len(Ea) > 8:
        lgia_sm = savgol_filter(lgia, max(5, min(11, len(Ea) // 2 * 2 - 1)), 3, mode="interp")
        dlg     = np.gradient(lgia_sm, Ea); adlg = np.abs(dlg)

        # ── Robust passive threshold (fixes ba=184/no-passive on real data) ──
        # OLD: thr = max(percentile_30, 0.80) — the hard floor of 0.80 fails when:
        #   (a) the passive plateau slope (0.68 dec/V) is below 0.80 but not
        #       consistently, due to noise breaking the flat run, OR
        #   (b) the percentile_30 itself is high (e.g. 6 dec/V for active-dominated
        #       curves), making thr=6 and nothing looks flat.
        #
        # NEW: Use TWO thresholds and take the union:
        #   1. Relative: thr_rel = percentile_10 * 4.0
        #      (4× the quietest 10% of the derivative — passive is much flatter
        #       than the active/transpassive wings even for gradual transitions)
        #   2. Absolute: thr_abs = 2.0 dec/V
        #      (passive plateau rarely exceeds 2 dec/V; active/transpassive do)
        # A point is "flat" if EITHER condition holds.
        # This reliably detects the passive plateau regardless of overall curve shape.
        p10 = np.percentile(adlg, 10)
        thr_rel = p10 * 4.0    # 4× quietest region
        thr_abs = 2.0           # absolute ceiling for passive slope
        # Use the SMALLER of the two (most selective that still finds flat regions)
        # but ensure it's at least as large as the relative threshold
        thr = max(thr_rel, min(thr_abs, np.percentile(adlg, 25)))
        flat = adlg < thr

        runs = [(k, list(g)) for k, g in groupby(enumerate(flat), key=lambda x: x[1]) if k]
        for _, ri in runs:
            idxs = [s[0] for s in ri]
            span = abs(Ea[idxs[-1]] - Ea[idxs[0]])
            if len(idxs) >= 4 and span > 0.05:
                ip_cand = float(np.median(np.abs(i[ano][si][idxs])))
                if ip_cand < float(np.max(np.abs(i[ano]))) * 0.7:
                    has_passive = True
                    Epass       = float(Ea[idxs[0]])
                    ip          = ip_cand
                    post = Ea > Ea[idxs[-1]]
                    if np.sum(post) > 3:
                        post_dlg = dlg[np.where(post)[0]]
                        post_E   = Ea[np.where(post)[0]]
                        # Transpassive: use relative threshold (3× local passive slope)
                        # OLD: post_dlg > 3.0 (absolute) — missed gentle transpassive
                        # Transpassive threshold: use fixed 1.5 dec/V minimum
                        # (was 3×thr which over-scales when thr is large)
                        trans_thr = max(thr, 1.5)
                        rising   = np.where(post_dlg > trans_thr)[0]
                        if len(rising) > 0:
                            has_trans = True
                            Etrans    = float(post_E[rising[0]])
                    break

    # ── Tafel window search ───────────────────────────────────────────────────
    base_mask = Ea > (Ecorr + CFG["anod_guard"])
    if np.sum(base_mask) < CFG["min_w_ano"]:
        base_mask = np.ones_like(Ea, bool)

    # CRITICAL FIX A: bound search to the ACTIVE dissolution zone only.
    # For Active-Passive curves the Tafel line must end at the active peak
    # (Epeak), NOT at Epass.  The old code used Epass as the upper bound,
    # but on steep active-passive curves Epass > Epeak and many of the
    # candidate windows span past the peak into the passive drop — their
    # mean slope is near-zero (high R², but beta >> 400 mV/dec) so they
    # are correctly killed by the hard beta gate.  However when Epeak is
    # absent (flat monotone rise straight into passivation, common on SS)
    # we should still bound by Epass to keep the search in the active region.
    E_upper = None
    if Epeak_detected is not None and Epeak_detected > Ecorr:
        E_upper = Epeak_detected        # tightest: stop at the active peak
    elif has_passive and Epass is not None:
        E_upper = Epass                 # fallback: stop at passivation onset
    else:
        # Last-resort fallback: derivative-drop bound.
        # When neither a distinct Epeak nor a flat passive plateau is found
        # (e.g. gradual monotone rise into passive with no local maximum),
        # find where the local derivative drops to < 40% of its peak value.
        # That transition marks the end of the steep active Tafel zone.
        # Without this, the window extends far into the passive/transpassive
        # region, inflating ba (e.g. producing ba=184 mV/dec on the image curve).
        _w_tmp = max(5, min(11, len(Ea)//2*2-1))
        _Y_tmp = savgol_filter(lgia, _w_tmp, 3, mode="interp")
        _dY_tmp = np.abs(np.gradient(_Y_tmp, Ea))
        # Only look within first 300 mV of anodic range for the active peak
        _near_mask = (Ea > Ecorr + CFG["anod_guard"]) & (Ea < Ecorr + 0.30)
        if np.sum(_near_mask) >= 4:
            _max_slope = float(np.max(_dY_tmp[_near_mask]))
            _drop_thr  = 0.40 * _max_slope
            # Find first SUSTAINED drop below 40% (2+ consecutive points)
            _cands = np.where(_dY_tmp < _drop_thr)[0]
            for _k in range(len(_cands) - 1):
                _idx = _cands[_k]
                if (_cands[_k+1] == _idx + 1
                        and Ea[_idx] > Ecorr + 0.020):
                    E_upper = float(Ea[_idx])
                    break

    if E_upper is not None:
        active_mask = base_mask & (Ea <= E_upper)
        if np.sum(active_mask) >= CFG["min_w_ano"]:
            base_mask = active_mask
        # else: too few points before peak/pass — keep full range, beta gate handles it

    Ex, Yx = Ea[base_mask], lgia[base_mask]
    if len(Ex) < CFG["min_w_ano"]:
        return dict(ba=0.060, has_passive=has_passive, Epass=Epass, ip=ip,
                    has_trans=has_trans, Etrans=Etrans, Epeak=Epeak_detected,
                    r2=0.0, E_an=Ea, lgi_an=lgia)

    w_sm = max(5, min(11, len(Ex) // 2 * 2 - 1))
    Y_sm = savgol_filter(Yx, w_sm, 3, mode="interp")
    dY   = np.gradient(Y_sm, Ex)
    d2Y  = np.gradient(dY, Ex)

    # ── Rising-only trim ──────────────────────────────────────────────────────
    # After the E_upper bound the data should be monotonically rising (active
    # dissolution), but noise in a long passive plateau can produce spurious
    # sign-changes that dilute the active slope estimate.
    # Trim any suffix where dY turns persistently negative (passive descent),
    # keeping only the contiguous rising front from Ecorr upward.
    # This was the root cause of ba=384 mV/dec in the image: the window
    # extended into the passive drop where apparent slope ≈ 0, i.e. ba >> 250.
    first_neg_run = None
    for _k in range(len(dY) - 2):
        if dY[_k] < 0 and dY[_k + 1] < 0 and dY[min(_k+2, len(dY)-1)] < 0:   # three consecutive negatives = descent
            first_neg_run = _k
            break
    if first_neg_run is not None and first_neg_run >= CFG["min_w_ano"]:
        Ex  = Ex[:first_neg_run]
        Yx  = Yx[:first_neg_run]
        dY  = dY[:first_neg_run]
        d2Y = d2Y[:first_neg_run]
        # Rerun sliding regression on trimmed data
        if len(Ex) < CFG["min_w_ano"]:
            Ex  = Ea[base_mask]; Yx  = lgia[base_mask]
            w_sm2 = max(5, min(11, len(Ex)//2*2-1))
            Y_sm2 = savgol_filter(Yx, w_sm2, 3, mode='interp')
            dY  = np.gradient(Y_sm2, Ex)
            d2Y = np.gradient(dY, Ex)

    # ── Anodic Tafel slope estimation ────────────────────────────────────────
    # The active window is already bounded by Epeak or Epass, so ALL points in
    # Ex belong to the active dissolution region. The task is simply to find
    # the STEEPEST linear sub-segment within Ex — that is the Tafel slope.
    #
    # Previous approach used proximity-to-Ecorr scoring which incorrectly
    # rewarded passive-transition windows near Ecorr over the true steep Tafel
    # region (producing ba=184 instead of ~95 mV/dec on the image data).
    #
    # New approach:
    #   • Sparse active zone (≤12 pts): use full-window OLS/Theil-Sen.
    #     Sub-window selection is unreliable with few noisy points.
    #   • Dense active zone (>12 pts): sliding window, scored purely by
    #     |slope| magnitude (steepest = most Tafel-like) and R².
    #     No proximity term — the E_upper bound already ensures we're in the
    #     correct zone.

    if len(Ex) <= 12:
        # ── Sparse path: fit all active points directly ───────────────────
        # With few points a sub-window just picks the flattest local segment.
        # Use the full window — Theil-Sen is robust against 1-2 noise outliers.
        sl_ts, b_ts = _theil_sen(Ex, Yx)
        sl_ols, b_ols = _theil_sen(Ex, Yx)          # same for OLS start
        sl_hb, b_hb = _huber_fit(Ex, Yx, sl_ts, b_ts)

        def _r2_full(sl, b):
            return r2_score(Yx, sl * Ex + b)
        cand = [(sl_ts, b_ts, _r2_full(sl_ts, b_ts)),
                (sl_hb, b_hb, _r2_full(sl_hb, b_hb))]
        sl_ref, b_ref, r2_win = max(cand, key=lambda t: t[2])

        # Ensure the slope is physically sensible
        if sl_ref <= 0 or abs(1.0 / sl_ref) > CFG["beta_max_a"] * 1.5:
            sl_ref = sl_ts
            b_ref  = b_ts
            r2_win = _r2_full(sl_ref, b_ref)

        ba = min(abs(1.0 / sl_ref), CFG["beta_max_a"]) if abs(sl_ref) > 1e-9 else 0.060
        win_e0, win_e1 = float(Ex[0]), float(Ex[-1])

    else:
        # ── Dense path: sliding window, steep-slope scoring ───────────────
        s_idx, e_idx, slope, intercept, R2 = _sliding_regress_full(
            Ex, Yx, min_len=CFG["min_w_ano"], max_len=20)
        if len(slope) == 0:
            return dict(ba=0.060, has_passive=has_passive, Epass=Epass, ip=ip,
                        has_trans=has_trans, Etrans=Etrans, Epeak=Epeak_detected,
                        r2=0.0, E_an=Ea, lgi_an=lgia)

        invm    = np.where(np.abs(slope) > 1e-12, 1.0 / np.abs(slope), np.inf)
        beta_ok = (slope > 0) & (invm > CFG["beta_min"]) & (invm < CFG["beta_max_a"])

        # Score by R² × window-length.
        # R² captures linearity (the true Tafel region is most linear).
        # Length rewards longer stable segments over short noisy ones.
        # No proximity or steepness bias — these caused wrong window
        # selection in both simple-active (steep bias picked mixed-control
        # zone near Ecorr) and active-passive (proximity bias picked passive
        # transition zone away from Ecorr).
        win_len_a    = (e_idx - s_idx).astype(float)
        len_bonus_a  = np.log1p(win_len_a) / np.log1p(25.0)  # normalised [0,1]
        score        = (R2 ** 2) * (0.6 + 0.4 * len_bonus_a)
        score[~beta_ok] = 0.0      # hard zero for unphysical slopes

        if np.all(score == 0.0):
            pos_mag = np.where(slope > 0, slope, 0.0)
            k_best  = int(np.argmax(pos_mag))
        else:
            k_best  = int(np.argmax(score))

        s0, s1 = int(s_idx[k_best]), int(e_idx[k_best])
        sl_ols, b_ols = float(slope[k_best]), float(intercept[k_best])

        sl_ts, b_ts = _theil_sen(Ex[s0:s1], Yx[s0:s1])
        sl_hb, b_hb = _huber_fit(Ex[s0:s1], Yx[s0:s1], sl_ols, b_ols)

        def _r2_line(sl, b):
            return r2_score(Yx[s0:s1], sl * Ex[s0:s1] + b)
        cand = [(sl_ts, b_ts, _r2_line(sl_ts, b_ts)),
                (sl_hb, b_hb, _r2_line(sl_hb, b_hb)),
                (sl_ols, b_ols, _r2_line(sl_ols, b_ols))]
        sl_ref, b_ref, r2_win = max(cand, key=lambda t: t[2])

        if sl_ref <= 0 or abs(1.0 / sl_ref) > CFG["beta_max_a"] * 1.5:
            sl_ref, b_ref = sl_ols, b_ols
            r2_win = _r2_line(sl_ref, b_ref)

        ba = min(abs(1.0 / sl_ref), CFG["beta_max_a"]) if abs(sl_ref) > 1e-9 else 0.060
        win_e0, win_e1 = float(Ex[s0]), float(Ex[s1 - 1])

    return dict(
        ba=ba, has_passive=has_passive, Epass=Epass, ip=ip,
        has_trans=has_trans, Etrans=Etrans, Epeak=Epeak_detected,
        r2=float(r2_win),
        E_an=Ea, lgi_an=lgia,
        slope_a=sl_ref, intercept_a=b_ref, win_a=(win_e0, win_e1)
    )

# ─────────────────────────────────────────────────────────────────────────────
# STAGE 4 — CLASSIFY CURVE TYPE
# ─────────────────────────────────────────────────────────────────────────────
def classify_curve(cat_res, an_res):
    hp = an_res["has_passive"]; ht = an_res["has_trans"]; hd = cat_res["has_diff"]
    if hp and ht:  return CT.F  if hd else CT.PT
    if hp:         return CT.P
    if hd:         return CT.AD
    return CT.A

# ─────────────────────────────────────────────────────────────────────────────
# TAFEL INTERSECTION (local lines)
# ─────────────────────────────────────────────────────────────────────────────
def tafel_intersection(cat_res, an_res):
    if ("slope_c" in cat_res and "intercept_c" in cat_res and
        "slope_a" in an_res  and "intercept_a" in an_res):
        mc = float(cat_res["slope_c"]); bc = float(cat_res["intercept_c"])
        ma = float(an_res["slope_a"]);  ba = float(an_res["intercept_a"])
        if abs(ma - mc) > 1e-12:
            E_star = (bc - ba) / (ma - mc)
            logI_star = ma * E_star + ba
            i_star = 10.0 ** logI_star
            return float(E_star), float(i_star), float(logI_star)
    return None

# ─────────────────────────────────────────────────────────────────────────────
# STAGE 5 — ASSEMBLE P0 & GLOBAL POLISH (LEANER)
# ─────────────────────────────────────────────────────────────────────────────
def _make_p0(Ecorr, cat, an, ct, E_max):
    ic, bc, ba, iL = cat["icorr"], cat["bc"], an["ba"], cat["iL"]
    Ep = an.get("Epeak") if an.get("Epeak") is not None else (an["Epass"] if an["has_passive"] and an["Epass"] is not None else E_max + 5.0)
    ip = an["ip"] if an["has_passive"] else ic * 0.01
    Et = an["Etrans"] if an["has_trans"] else E_max + 5.0
    it = ip * 0.5 if an["has_trans"] else ic * 0.001
    return np.array([Ecorr, ic, ba, bc, Ep, 0.010, ip, Et, 0.015, it, iL])

def _build_bounds(Ecorr, cat, an, ct, E_min, E_max, E_span):
    """
    Build parameter bounds for the global optimizer.

    Inspired by busbar_tafel_fit.py:
    - ba/bc: [max(0.020, local×0.50), min(0.500, local×2.0)]
      Absolute floor (20 mV/dec) + ceiling (500 mV/dec) prevent runaway while
      remaining wide enough to correct a ±50% error in the local Tafel estimate.
    - iL: lo = ic×5 (must exceed icorr to be physically meaningful)
          hi = i_max×50 (cap at 50× measured maximum current)
    - icorr: ×1e-4 / ×1e4 relative to local estimate (generous range)
    """
    ic      = max(cat["icorr"], 1e-14)
    iL_est  = max(cat["iL"],    1e-10)
    ba_fit  = float(an["ba"])
    bc_fit  = float(cat["bc"])

    # Tafel slopes: tight relative band + absolute physical limits.
    # Upper caps lowered to match CFG beta_max values — prevents the global
    # optimizer from finding unphysical solutions (ba>250mV or bc>280mV)
    # even if the local Tafel estimate happened to be inflated.
    ba_lo = max(ba_fit * 0.50, 0.020)
    ba_hi = min(ba_fit * 2.00, CFG["beta_max_a"])   # hard ceiling 250 mV/dec
    bc_lo = max(bc_fit * 0.50, 0.020)
    bc_hi = min(bc_fit * 2.00, CFG["beta_max_c"])   # hard ceiling 280 mV/dec

    # iL: must be physically larger than icorr; capped by measured data range
    i_max = max(float(np.max(np.abs(cat.get("iL", ic * 100)))), ic * 10)
    iL_lo = max(ic * 5.0,   1e-13)
    iL_hi = min(i_max * 50, 1.0)

    lo = np.array([E_min,
                   max(ic * 1e-4, 1e-15),     # icorr lo
                   ba_lo, bc_lo,
                   Ecorr + 0.005, 0.001,       # Epass, k_pass
                   max(ic * 1e-5, 1e-16),      # ip lo
                   Ecorr + 0.05 * E_span, 0.001,  # Etrans, k_trans
                   max(ic * 1e-5, 1e-16),      # itrans lo
                   iL_lo])
    hi = np.array([E_max,
                   min(ic * 1e4, 1.0),         # icorr hi
                   ba_hi, bc_hi,
                   E_max, 0.150,               # Epass, k_pass
                   min(ic * 1e3, 1.0),         # ip hi
                   E_max + 0.1, 0.150,         # Etrans, k_trans
                   min(ic * 1e5, 10.0),        # itrans hi
                   iL_hi])
    lo = np.minimum(lo, hi - 1e-12)
    return lo, hi

LOG_IDX = {1, 6, 9, 10}

def _pack(p, fidx):
    return np.array([np.log10(max(p[j], TINY)) if j in LOG_IDX else p[j] for j in fidx])

def _unpack(x, fidx, p_base, lo=None, hi=None):
    p = p_base.copy()
    for k, j in enumerate(fidx):
        val = 10.0 ** x[k] if j in LOG_IDX else x[k]
        if lo is not None: val = float(np.clip(val, lo[j], hi[j]))
        p[j] = val
    return p

def _pbounds(lo, hi, fidx):
    return [(np.log10(max(lo[j], TINY)), np.log10(max(hi[j], TINY)))
            if j in LOG_IDX else (lo[j], hi[j]) for j in fidx]

def _section_weights(E, i, Ecorr_est, Epass_est=None):
    """
    Section-balanced weighting for the global objective function.

    Motivation (from busbar analysis):
    - For Active-Passive curves the passive plateau may hold 60%+ of all points.
      Uniform log-residual weighting then forces the optimizer to over-fit the
      trivially-flat passive region at the expense of the kinetically important
      cathodic and active branches.
    - Strategy: each electrochemical region (cathodic, active, passive) contributes
      roughly equally to the total objective, regardless of point density.
    - Additional Ecorr-vicinity boost (×3 within 60 mV) sharpens the corrosion
      potential and exchange current density.
    - Passive plateau is further down-weighted (×0.5) because it is
      well-described by a single parameter (ip) and needs little optimizer effort.
    """
    n = len(E)
    cat_m  = E < Ecorr_est
    if Epass_est is not None:
        act_m  = (E >= Ecorr_est) & (E < Epass_est)
        pass_m = E >= Epass_est
    else:
        act_m  = (E >= Ecorr_est) & (E < Ecorr_est + 0.10)
        pass_m = E >= (Ecorr_est + 0.10)

    n_cat  = max(int(cat_m.sum()),  1)
    n_act  = max(int(act_m.sum()),  1)
    n_pass = max(int(pass_m.sum()), 1)

    w = np.ones(n, float)
    # Equal section contribution
    w[cat_m]  *= n / (3.0 * n_cat)
    w[act_m]  *= n / (3.0 * n_act)  * 1.5   # active Tafel: slightly boosted
    w[pass_m] *= n / (3.0 * n_pass) * 0.50  # passive plateau: easy, reduce weight

    # Ecorr vicinity boost (×3 within 60 mV) — sharpens icorr and Ecorr
    w *= 1.0 + 2.0 * np.exp(-np.abs(E - Ecorr_est) / 0.060)
    w /= w.mean()
    return w

def global_polish(E, i, p0, ct, lo, hi):
    """
    Global parameter optimisation: DE → L-BFGS-B → Nelder-Mead → Powell.

    Improvements over previous version (inspired by busbar_tafel_fit.py):
    - 4-stage cascade: adds Nelder-Mead (adaptive) between L-BFGS-B and Powell,
      which handles non-smooth basins that L-BFGS-B misses.
    - Tighter tolerances: ftol=1e-15, gtol=1e-13 throughout.
    - Larger DE population: popsize=max(15, nf×3), maxiter=max(400, nf×60).
    - Section-balanced objective weights (see _section_weights) replace the
      old Ecorr-only Gaussian weight that over-concentrated on the mixed-control
      zone and starved the cathodic/passive regions of gradient signal.
    - Bounds enforced via clipping inside _unpack for gradient-free methods
      (Nelder-Mead, Powell) that don't natively respect bounds.
    - ba/bc runaway detection: flags fits where Tafel slopes hit the 500 mV/dec
      upper bound (unphysical — indicates no real linear Tafel region detected).
    """
    ld   = slog(i)
    fidx = CT.idx(ct)
    bnds = _pbounds(lo, hi, fidx)
    n, nf = len(E), len(fidx)

    Ecorr_p0 = float(p0[0])
    Epass_p0 = float(p0[4]) if ct in CT.PASS else None
    w_base   = _section_weights(E, i, Ecorr_p0, Epass_p0)

    def obj(x):
        p = _unpack(x, fidx, p0.copy(), lo, hi)
        try:
            pred = pol_model(E, p, ct)
            return float(np.sum(w_base * (ld - slog(pred)) ** 2))
        except Exception:
            return 1e30

    best_x = _pack(p0, fidx)
    best_val = obj(best_x)

    def update(x):
        nonlocal best_x, best_val
        v = obj(x)
        if v < best_val - 1e-12:
            best_x, best_val = x.copy(), v
        return v

    # ── Stage 1: Differential Evolution (global search) ───────────────────
    # Larger population and more iterations than before (busbar: popsize≥15)
    ps = max(15, nf * 3)
    mi = max(400, nf * 60)
    for seed, strat in [(42, "best1bin"), (7, "currenttobest1bin")]:
        try:
            res = differential_evolution(
                obj, bnds, seed=seed, maxiter=mi, popsize=ps,
                tol=1e-12, mutation=(0.5, 1.9), recombination=0.90,
                polish=False, workers=1, strategy=strat)
            update(res.x)
        except Exception:
            pass

    # ── Stage 2: L-BFGS-B (gradient-based refinement) ────────────────────
    try:
        r = minimize(obj, best_x, method="L-BFGS-B", bounds=bnds,
                     options={"maxiter": 25000, "ftol": 1e-15, "gtol": 1e-13})
        update(r.x)
    except Exception:
        pass

    # ── Stage 3: Nelder-Mead adaptive (non-smooth basin escape) ──────────
    # From busbar: handles parameter correlations that trip up L-BFGS-B
    try:
        r = minimize(obj, best_x, method="Nelder-Mead",
                     options={"maxiter": 20000, "xatol": 1e-12,
                              "fatol": 1e-14, "adaptive": True})
        update(r.x)
    except Exception:
        pass

    # ── Stage 4: Powell (final coordinate-wise polish) ────────────────────
    try:
        r = minimize(obj, best_x, method="Powell",
                     options={"maxiter": 15000, "xtol": 1e-12, "ftol": 1e-14})
        update(r.x)
    except Exception:
        pass

    best_p = _unpack(best_x, fidx, p0.copy(), lo, hi)
    log_p  = slog(pol_model(E, best_p, ct))
    sse    = float(np.sum((ld - log_p) ** 2))
    r2     = r2_score(ld, log_p)
    aic    = aicc(n, nf, sse)
    return best_p, r2, aic, sse

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────
COL_SIG = [
    (r"we.*potential|ewe|potential/v|e/v|^e$|e \(v\)|e_v|^vf$", r"we.*current|<i>/ma|i/ma|current/a|i/a|^i$|i \(a\)|i_a|^im$", "A"),
    (r"potential|volt|^e$", r"current.*ma|ima", "mA"),
]
UNIT_PAT = {r"\(a\)|_a$|/a$|a/cm²?": 1.0, r"\(ma\)|_ma$|/ma$|ma/cm²?": 1e-3, r"\(ua\)|_ua$|/ua$|ua/cm²?": 1e-6}

def _auto_cols(df):
    num = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    cl  = {c: c.lower().strip() for c in df.columns}
    for Ep, Ip, uf in COL_SIG:
        em = [c for c, v in cl.items() if re.search(Ep, v) and c in num]
        im = [c for c, v in cl.items() if re.search(Ip, v) and c in num and c not in em]
        if em and im:
            ec = sorted(em, key=lambda c: 0 if "we" in cl[c] else 1)[0]
            ic = im[0]
            f  = 1e-3 if uf == "mA" else 1.0
            for pat, fv in UNIT_PAT.items():
                if re.search(pat, cl[ic]): f = fv; break
            return ec, ic, f
    if len(num) >= 2: return num[0], num[1], 1.0
    raise ValueError("Cannot detect E/i columns automatically.")

def load_file(uploaded):
    name = uploaded.name.lower()
    if name.endswith((".xlsx", ".xls")): return pd.read_excel(uploaded)
    content = uploaded.getvalue().decode("utf-8", errors="replace")
    for sep in ["\t", ";", ",", r"\s+"]:
        try:
            df = pd.read_csv(io.StringIO(content), sep=sep, engine="python", comment="#")
            if df.shape[1] >= 2 and df.shape[0] > 4: return df.dropna(axis=1, how="all")
        except: continue
    raise ValueError("Cannot parse file.")

# ─────────────────────────────────────────────────────────────────────────────
# PUBLICATION FIGURE (LOCAL-LINE TAFEL OVERLAYS)
# FIX: Linear-scale panel uses E_corr-local ylim instead of global p95
# ─────────────────────────────────────────────────────────────────────────────
PLT_RC = {
    "font.family": "DejaVu Sans", "font.size": 10,
    "axes.titlesize": 11, "axes.titleweight": "bold",
    "axes.labelsize": 10, "axes.linewidth": 0.9,
    "xtick.direction": "in", "ytick.direction": "in",
    "xtick.major.size": 4, "ytick.major.size": 4,
    "xtick.minor.size": 2.5, "ytick.minor.size": 2.5,
    "xtick.major.width": 0.8, "ytick.major.width": 0.8,
    "legend.fontsize": 8, "legend.framealpha": 0.92,
    "legend.edgecolor": "#cccccc", "grid.color": "#e0e0e0",
    "grid.linewidth": 0.6, "figure.facecolor": "white",
    "axes.facecolor": "#fafbff",
}

def make_figure(E, i_obs, best_p, ct, sample_name, cat_res, an_res,
                Ecorr, taf=None, extend_tafel=True, use_tafel_icorr=True,
                show_regions=True, dpi=150):
    """
    Publication-grade figure with 2-row layout:
      Row 1: Full Evans Diagram (spanning all columns)
      Row 2: Cathodic Tafel zoom | Anodic/Active zoom | Passive+Trans zoom | Residuals

    Tafel lines are drawn ONLY within their fitted linear windows (solid thick),
    then extended as thin dashed lines toward the Ecorr intersection.
    No model-partial fallback lines — if local fit is absent, nothing is drawn.
    The Ecorr and icorr are always shown as the Tafel-line intersection point.
    """
    Ecorr_fit = float(best_p[0])
    ba    = max(float(best_p[2]), 1e-9)
    bc    = max(float(best_p[3]), 1e-9)
    icorr_model = float(best_p[1])

    E_lo, E_hi = float(E.min()), float(E.max())
    span = max(E_hi - E_lo, 1e-6)
    n_dense = int(np.clip(200 * span, 800, 2500))
    E_dense = np.linspace(E_lo, E_hi, n_dense)

    i_dense  = pol_model(E_dense, best_p, ct)
    i_fit_E  = pol_model(E, best_p, ct)
    log_obs  = slog(i_obs)
    log_den  = slog(i_dense)
    log_fitE = slog(i_fit_E)
    residuals = log_obs - log_fitE
    r2v  = r2_score(log_obs, log_fitE)
    rmse = float(np.sqrt(np.mean(residuals**2)))

    fin  = log_obs[np.isfinite(log_obs)]
    y_lo = float(np.nanmin(fin)) - 0.2

    if ct in CT.PASS:
        Epass_4y = float(best_p[4])
        act_mask = (E >= Ecorr_fit - 0.02) & (E <= Epass_4y + 0.05)
        log_act = slog(i_obs[act_mask])
        log_act = log_act[np.isfinite(log_act)]
        active_peak_log = float(np.max(log_act)) if len(log_act) > 0 else np.log10(max(icorr_model, TINY))
        logIp_val = float(np.log10(max(float(best_p[6]), TINY)))
        y_hi = max(active_peak_log + 1.5, logIp_val + 2.0)
    else:
        y_hi = float(np.percentile(fin, 90)) + 1.0
    y_hi = min(y_hi, float(np.percentile(fin, 99)) + 1.8)

    # ── Tafel intersection (Ecorr, icorr from linear intersection) ──────────
    icorr_display, ecorr_display = icorr_model, Ecorr_fit
    if taf is not None and use_tafel_icorr:
        E_tafel, i_tafel, logI_tafel = taf
        if E_lo - 0.2*span <= E_tafel <= E_hi + 0.2*span and np.isfinite(i_tafel) and i_tafel > 0:
            icorr_display, ecorr_display = i_tafel, E_tafel
    logIc_disp = np.log10(max(icorr_display, TINY))

    # ── Active-zone upper bound for anodic line drawing ──────────────────────
    # Prefer Epeak, fall back to Epass from model
    active_upper = None
    if an_res.get("Epeak") is not None:
        active_upper = float(an_res["Epeak"])
    elif ct in CT.PASS:
        active_upper = float(best_p[4])

    # ── Helper: draw a Tafel line segment + optional thin extension ──────────
    def _tafel_line(ax, slope, intercept, win_lo, win_hi,
                    color, lw_main=2.2, lw_ext=1.4, alpha_ext=0.45,
                    label=None, extend_left=None, extend_right=None, clip_lo=None, clip_hi=None):
        """
        Draw the Tafel line within [win_lo, win_hi] as a solid line (main region),
        then optionally extend to extend_left / extend_right as thin dashed line.
        clip_lo / clip_hi limit the drawn range to plot axes.
        """
        lo = max(win_lo, clip_lo if clip_lo is not None else win_lo)
        hi = min(win_hi, clip_hi if clip_hi is not None else win_hi)
        if hi <= lo:
            return
        Eseg = np.linspace(lo, hi, 120)
        ax.plot(Eseg, slope*Eseg + intercept, "-", color=color, lw=lw_main,
                zorder=7, label=label)
        # Extension toward Ecorr (left for cathodic, right for anodic)
        if extend_left is not None and extend_left < lo:
            el = max(extend_left, clip_lo if clip_lo is not None else extend_left)
            if el < lo:
                Eext = np.linspace(el, lo, 80)
                ax.plot(Eext, slope*Eext + intercept, "--", color=color,
                        lw=lw_ext, alpha=alpha_ext, zorder=6)
        if extend_right is not None and extend_right > hi:
            er = min(extend_right, clip_hi if clip_hi is not None else extend_right)
            if er > hi:
                Eext = np.linspace(hi, er, 80)
                ax.plot(Eext, slope*Eext + intercept, "--", color=color,
                        lw=lw_ext, alpha=alpha_ext, zorder=6)

    with plt.rc_context(PLT_RC):
        # ── 2-row layout: Evans (top, full width) + 4 region panels (bottom) ─
        fig = plt.figure(figsize=(18, 11), dpi=dpi)
        gs  = GridSpec(2, 4, figure=fig,
                       hspace=0.48, wspace=0.38,
                       left=0.06, right=0.98, top=0.93, bottom=0.08)
        ax_ev   = fig.add_subplot(gs[0, :])       # full Evans diagram
        ax_cat  = fig.add_subplot(gs[1, 0])        # cathodic Tafel zoom
        ax_ano  = fig.add_subplot(gs[1, 1])        # anodic active zoom
        ax_pass = fig.add_subplot(gs[1, 2])        # passive (+ transpassive)
        ax_res  = fig.add_subplot(gs[1, 3])        # residuals

        # ══════════════════════════════════════════════════════════════════════
        # PANEL A — Full Evans Diagram
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_ev

        # Region shading
        if show_regions:
            def vband(ax, e0, e1, key, lbl):
                c, a = REGION_COLORS[key]
                e0c = float(np.clip(e0, E_lo, E_hi))
                e1c = float(np.clip(e1, E_lo, E_hi))
                if e1c > e0c:
                    ax.axvspan(e0c, e1c, color=c, alpha=a, lw=0, label=lbl, zorder=1)
            vband(ax, E_lo, Ecorr_fit, "cathodic", "Cathodic")
            if ct in CT.SIMPLE:
                vband(ax, Ecorr_fit, E_hi, "active", "Anodic (active)")
            elif ct in CT.PASS:
                Ep = float(best_p[4])
                Et = float(best_p[7]) if ct in CT.TRANS else E_hi + 1
                vband(ax, Ecorr_fit, min(Ep, E_hi),    "active",       "Active dissolution")
                vband(ax, min(Ep, E_hi), min(Et, E_hi), "passive",     "Passive region")
                if ct in CT.TRANS and Et < E_hi:
                    vband(ax, Et, E_hi, "transpassive", "Transpassive / pitting")

        # Data + global fit
        ax.scatter(E, log_obs, s=12, color="#4a7fa8", alpha=0.55,
                   zorder=2, label="Experimental data", linewidths=0, rasterized=True)
        ax.plot(E_dense, log_den, color="#1a3a5c", lw=2.0, zorder=5,
                label=f"Global fit (R²={r2v:.5f})")

        # ── Cathodic Tafel line (solid in window, dashed extension to Ecorr) ─
        if "slope_c" in cat_res and "win_c" in cat_res:
            sc, ic_int = cat_res["slope_c"], cat_res["intercept_c"]
            wc0, wc1 = float(cat_res["win_c"][0]), float(cat_res["win_c"][1])
            bc_lbl = f"βc = {min(abs(1/sc), CFG['beta_max_c'])*1000:.0f} mV/dec"
            ext_r = ecorr_display if extend_tafel else None
            _tafel_line(ax, sc, ic_int, wc0, wc1, "#8e44ad",
                        label=bc_lbl, extend_right=ext_r,
                        clip_lo=E_lo, clip_hi=E_hi)

        # ── Anodic Tafel line (solid in active window, dashed extension to Ecorr) ─
        if "slope_a" in an_res and "win_a" in an_res:
            sa, ia_int = an_res["slope_a"], an_res["intercept_a"]
            wa0 = float(an_res["win_a"][0])
            # Clip right edge of anodic window to active zone (not passive!)
            wa1 = float(an_res["win_a"][1])
            if active_upper is not None:
                wa1 = min(wa1, active_upper)
            if wa1 <= wa0:
                wa1 = wa0 + 0.010   # ensure at least 10 mV visible
            ba_lbl = f"βa = {min(abs(1/sa), CFG['beta_max_a'])*1000:.0f} mV/dec"
            ext_l = ecorr_display if extend_tafel else None
            _tafel_line(ax, sa, ia_int, wa0, wa1, "#e67e22",
                        label=ba_lbl, extend_left=ext_l,
                        clip_lo=E_lo, clip_hi=E_hi)

        # i_pass horizontal line
        if ct in CT.PASS:
            ip_val = float(best_p[6])
            ax.axhline(np.log10(max(ip_val, TINY)), color="#27ae60",
                       ls=":", lw=1.2, alpha=0.80, zorder=3,
                       label=f"i_pass = {ip_val:.2e} A/cm²")

        # Intersection marker + drop lines
        ax.plot(ecorr_display, logIc_disp, "x", color="#e84393", ms=12, mew=2.5, zorder=9)
        ax.plot([ecorr_display]*2, [y_lo, logIc_disp], ":", color="#e84393", lw=1.2, alpha=0.9)
        ax.plot([E_lo, ecorr_display], [logIc_disp]*2,  ":", color="#e84393", lw=1.2, alpha=0.9)

        y_span = y_hi - y_lo
        ax.annotate(f"Eᶜᵒʳʳ = {ecorr_display:.4f} V",
                    xy=(ecorr_display, y_lo + 0.04*y_span),
                    fontsize=8.5, color="#e84393", fontweight="bold")
        ax.annotate(f"iᶜᵒʳʳ = {icorr_display:.2e} A/cm²",
                    xy=(E_lo + 0.01*span, logIc_disp + 0.03*y_span),
                    fontsize=8.5, color="#e84393", fontweight="bold")

        ax.set_xlim(E_lo, E_hi); ax.set_ylim(y_lo, y_hi)
        ax.set_xlabel("E vs. Reference (V)")
        ax.set_ylabel("log₁₀ |i| (A cm⁻²)")
        ax.set_title(f"Evans Diagram — {sample_name}")
        ax.xaxis.set_minor_locator(AutoMinorLocator(5)); ax.yaxis.set_minor_locator(AutoMinorLocator(5))
        ax.tick_params(which="both", top=True, right=True)
        ax.grid(True, which="major", ls="--", alpha=0.45)
        ax.grid(True, which="minor", ls=":", alpha=0.18)
        ax.legend(loc="lower right", ncol=5, fontsize=7.5, framealpha=0.95, edgecolor="#cccccc")
        r2c = "#27ae60" if r2v > 0.99 else "#e67e22" if r2v > 0.95 else "#e84393"
        ax.text(0.01, 0.97,
                f"R²={r2v:.5f}  RMSE={rmse:.4f}  Model: {CT.name(ct)}",
                transform=ax.transAxes, fontsize=8.5, color=r2c, fontweight="bold", va="top",
                bbox=dict(fc="white", ec=r2c, alpha=0.88, pad=3, boxstyle="round,pad=0.3"))

        # ══════════════════════════════════════════════════════════════════════
        # PANEL B — Cathodic Tafel region (zoomed)
        # Shows the cathodic data with the fitted linear region highlighted
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_cat
        if "E_cat" in cat_res:
            Ec_arr = cat_res["E_cat"]; lgi_c = cat_res["lgi_cat"]
            ax.scatter(Ec_arr, lgi_c, s=20, color="#6baed6", alpha=0.75,
                       zorder=2, label="Cathodic data", linewidths=0, rasterized=True)
            # Highlight the fitted window
            if "win_c" in cat_res:
                wc0, wc1 = float(cat_res["win_c"][0]), float(cat_res["win_c"][1])
                win_mask = (Ec_arr >= wc0 - 0.002) & (Ec_arr <= wc1 + 0.002)
                if win_mask.sum() > 0:
                    ax.scatter(Ec_arr[win_mask], lgi_c[win_mask],
                               s=40, color="#2c3e8c", alpha=0.95, zorder=4,
                               label="Tafel window", linewidths=0)
            # Draw the fitted line over the window + extension to Ecorr
            if "slope_c" in cat_res:
                sc, ic_int = cat_res["slope_c"], cat_res["intercept_c"]
                wc0, wc1 = float(cat_res["win_c"][0]), float(cat_res["win_c"][1])
                # Solid line in window
                E_win = np.linspace(wc0, wc1, 100)
                ax.plot(E_win, sc*E_win + ic_int, "-", color="#8e44ad", lw=2.2, zorder=5,
                        label=f"βc = {min(abs(1/sc), CFG['beta_max_c'])*1000:.0f} mV/dec")
                # Dashed extension to Ecorr
                if extend_tafel and wc1 < ecorr_display:
                    E_ext = np.linspace(wc1, ecorr_display, 80)
                    ax.plot(E_ext, sc*E_ext + ic_int, "--", color="#8e44ad",
                            lw=1.4, alpha=0.50, zorder=4)
            ax.axvline(ecorr_display, color="#e84393", ls="--", lw=1.0, alpha=0.7)
            ax.axhline(logIc_disp, color="#e84393", ls=":", lw=0.9, alpha=0.7)
            # Global model cathodic curve for comparison
            cat_dense_m = E_dense <= Ecorr_fit + 0.01
            ax.plot(E_dense[cat_dense_m], log_den[cat_dense_m],
                    color="#1a3a5c", lw=1.5, alpha=0.60, zorder=3, ls="-",
                    label="Global model")
            # Axis limits: focus on the cathodic Tafel region
            xlim_c = (float(Ec_arr.min()) - 0.01, Ecorr_fit + 0.02)
            ylim_c_lo = float(np.nanmin(lgi_c)) - 0.1
            ylim_c_hi = float(np.nanmax(lgi_c)) + 0.2
            ax.set_xlim(xlim_c); ax.set_ylim(ylim_c_lo, ylim_c_hi)
        ax.set_xlabel("E (V)"); ax.set_ylabel("log₁₀ |i|")
        ax.set_title("Cathodic Tafel Region")
        ax.xaxis.set_minor_locator(AutoMinorLocator(4)); ax.yaxis.set_minor_locator(AutoMinorLocator(4))
        ax.tick_params(which="both", top=True, right=True)
        ax.grid(True, which="major", ls="--", alpha=0.4)
        ax.legend(fontsize=7.5)

        # ══════════════════════════════════════════════════════════════════════
        # PANEL C — Anodic active dissolution region (zoomed)
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_ano
        if "E_an" in an_res:
            Ea_arr = an_res["E_an"]; lgi_a = an_res["lgi_an"]
            # Determine active zone upper limit
            act_up = active_upper if active_upper is not None else float(Ea_arr.max())
            act_mask_an = Ea_arr <= act_up + 0.010
            Ea_act = Ea_arr[act_mask_an]; lgi_act = lgi_a[act_mask_an]
            ax.scatter(Ea_act, lgi_act, s=20, color="#fd8d3c", alpha=0.75,
                       zorder=2, label="Active data", linewidths=0, rasterized=True)
            # Highlight the fitted window
            if "win_a" in an_res:
                wa0 = float(an_res["win_a"][0])
                wa1 = min(float(an_res["win_a"][1]), act_up)
                win_mask_a = (Ea_act >= wa0 - 0.002) & (Ea_act <= wa1 + 0.002)
                if win_mask_a.sum() > 0:
                    ax.scatter(Ea_act[win_mask_a], lgi_act[win_mask_a],
                               s=45, color="#c0390b", alpha=0.95, zorder=4,
                               label="Tafel window", linewidths=0)
            # Draw the fitted Tafel line
            if "slope_a" in an_res and "win_a" in an_res:
                sa, ia_int = an_res["slope_a"], an_res["intercept_a"]
                wa0 = float(an_res["win_a"][0])
                wa1 = min(float(an_res["win_a"][1]), act_up)
                if wa1 > wa0:
                    E_win_a = np.linspace(wa0, wa1, 100)
                    ax.plot(E_win_a, sa*E_win_a + ia_int, "-", color="#e67e22", lw=2.2, zorder=5,
                            label=f"βa = {min(abs(1/sa), CFG['beta_max_a'])*1000:.0f} mV/dec")
                    # Dashed extension to Ecorr
                    if extend_tafel and ecorr_display < wa0:
                        E_ext_a = np.linspace(ecorr_display, wa0, 80)
                        ax.plot(E_ext_a, sa*E_ext_a + ia_int, "--", color="#e67e22",
                                lw=1.4, alpha=0.50, zorder=4)
            ax.axvline(ecorr_display, color="#e84393", ls="--", lw=1.0, alpha=0.7)
            ax.axhline(logIc_disp, color="#e84393", ls=":", lw=0.9, alpha=0.7)
            # Global model anodic
            ano_dense_m = (E_dense >= Ecorr_fit - 0.01) & (E_dense <= act_up + 0.02)
            ax.plot(E_dense[ano_dense_m], log_den[ano_dense_m],
                    color="#1a3a5c", lw=1.5, alpha=0.60, zorder=3, label="Global model")
            # Axis limits: tight around active zone
            xlim_a = (Ecorr_fit - 0.02, act_up + 0.02)
            if len(lgi_act) > 0:
                ylim_a_lo = float(np.nanmin(lgi_act)) - 0.1
                ylim_a_hi = float(np.nanmax(lgi_act)) + 0.3
                ax.set_xlim(xlim_a); ax.set_ylim(ylim_a_lo, ylim_a_hi)
        ax.set_xlabel("E (V)"); ax.set_ylabel("log₁₀ |i|")
        ax.set_title("Anodic Active Region")
        ax.xaxis.set_minor_locator(AutoMinorLocator(4)); ax.yaxis.set_minor_locator(AutoMinorLocator(4))
        ax.tick_params(which="both", top=True, right=True)
        ax.grid(True, which="major", ls="--", alpha=0.4)
        ax.legend(fontsize=7.5)

        # ══════════════════════════════════════════════════════════════════════
        # PANEL D — Passive + Transpassive region (or Linear scale for active curves)
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_pass
        if ct in CT.PASS:
            # Show passive plateau and transpassive, with model overlay
            Ep_fit = float(best_p[4])
            pass_mask_E = E >= Ep_fit - 0.02
            E_pass_data  = E[pass_mask_E]; i_pass_data  = i_obs[pass_mask_E]
            log_pass_data = slog(i_pass_data)
            ax.scatter(E_pass_data, log_pass_data, s=14, color="#74c476", alpha=0.70,
                       zorder=2, label="Passive / Trans data", linewidths=0, rasterized=True)
            # Global model over this range
            pass_dense_m = E_dense >= Ep_fit - 0.03
            ax.plot(E_dense[pass_dense_m], log_den[pass_dense_m],
                    color="#1a3a5c", lw=2.0, zorder=5, label="Global model")
            # i_pass line
            ip_val = float(best_p[6])
            ax.axhline(np.log10(max(ip_val, TINY)), color="#27ae60",
                       ls="--", lw=1.4, alpha=0.85, label=f"i_pass={ip_val:.2e}")
            # Epass and Etrans markers
            ax.axvline(Ep_fit, color="#27ae60", ls="-.", lw=1.0,
                       label=f"E_pass={Ep_fit:.3f}V")
            if ct in CT.TRANS:
                Et_fit = float(best_p[7])
                if E_lo <= Et_fit <= E_hi:
                    ax.axvline(Et_fit, color="#9e9ac8", ls="-.", lw=1.0,
                               label=f"E_trans={Et_fit:.3f}V")
            # Annotate passive current
            ax.annotate(f"ip = {ip_val:.2e}", xy=(Ep_fit + 0.01, np.log10(max(ip_val,TINY)) + 0.05),
                        fontsize=8, color="#27ae60")
            xlim_p = (Ep_fit - 0.03, E_hi)
            if len(log_pass_data) > 0:
                ylim_p_lo = float(np.nanmin(log_pass_data)) - 0.1
                ylim_p_hi = float(np.nanmax(log_pass_data)) + 0.3
                ax.set_xlim(xlim_p); ax.set_ylim(ylim_p_lo, ylim_p_hi)
            ax.set_title("Passive + Transpassive Region")
        else:
            # For non-passive curves: show linear i vs E (Stern plot)
            local_m = np.abs(E - ecorr_display) <= 0.150
            i_ref = float(np.percentile(np.abs(i_obs[local_m]), 95)) if local_m.sum() >= 4 else float(np.percentile(np.abs(i_obs), 80))
            uscale = 1e9 if i_ref < 1e-6 else 1e6 if i_ref < 1e-3 else 1e3
            ulbl   = "nA/cm²" if i_ref < 1e-6 else "μA/cm²" if i_ref < 1e-3 else "mA/cm²"
            ylim_l = max(i_ref * uscale * 1.4, 1e-12)
            ax.scatter(E, i_obs*uscale, s=9, color="#4a7fa8", alpha=0.6, label="Data", linewidths=0, rasterized=True)
            ax.plot(E_dense, np.clip(i_dense*uscale, -ylim_l*3, ylim_l*3), color="#1a3a5c", lw=2, label="Fit")
            ax.axhline(0, color="#888", lw=0.7); ax.axvline(ecorr_display, color="#e84393", ls="--", lw=1.0)
            ax.set_xlim(E_lo, E_hi); ax.set_ylim(-ylim_l, ylim_l)
            ax.set_ylabel(f"i ({ulbl})")
            ax.set_title("Linear Scale (Stern plot)")
        ax.set_xlabel("E (V)")
        ax.xaxis.set_minor_locator(AutoMinorLocator(4)); ax.yaxis.set_minor_locator(AutoMinorLocator(4))
        ax.tick_params(which="both", top=True, right=True)
        ax.grid(True, which="major", ls="--", alpha=0.4)
        ax.legend(fontsize=7.5)

        # ══════════════════════════════════════════════════════════════════════
        # PANEL E — Residuals (Δ log|i| vs E)
        # ══════════════════════════════════════════════════════════════════════
        ax = ax_res
        ax.fill_between([E_lo, E_hi], -0.1, 0.1, color="#e84393", alpha=0.07, zorder=1)
        ax.scatter(E, residuals, s=10, color="#2e86de", alpha=0.65, zorder=3,
                   linewidths=0, rasterized=True)
        ax.axhline(0,     color="#333",    lw=0.9, zorder=2)
        ax.axhline( 0.1,  color="#e84393", ls=":", lw=1.0, alpha=0.7)
        ax.axhline(-0.1,  color="#e84393", ls=":", lw=1.0, alpha=0.7, label="±0.1 log")
        ax.axvline(ecorr_display, color="#e84393", ls="--", lw=0.9, alpha=0.6)
        # Region delimiters in residual plot
        if ct in CT.PASS:
            ax.axvline(float(best_p[4]), color="#27ae60", ls="-.", lw=0.8, alpha=0.5)
            if ct in CT.TRANS:
                ax.axvline(float(best_p[7]), color="#9e9ac8", ls="-.", lw=0.8, alpha=0.5)
        ax.set_xlim(E_lo, E_hi)
        ax.set_xlabel("E (V)"); ax.set_ylabel("Δ log₁₀ |i|")
        ax.set_title(f"Residuals   R²={r2v:.5f}")
        ax.xaxis.set_minor_locator(AutoMinorLocator(4)); ax.yaxis.set_minor_locator(AutoMinorLocator(4))
        ax.tick_params(which="both", top=True, right=True)
        ax.grid(True, which="major", ls="--", alpha=0.4)
        ax.legend(fontsize=8)

        fig.suptitle("Polarisation Curve Analysis", fontsize=12,
                     fontweight="bold", color="#1a3a5c", y=0.98)

    return fig, r2v, rmse, icorr_display, ecorr_display

# ─────────────────────────────────────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def export_excel(results_list):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Results"
    H_FILL = PatternFill("solid", fgColor="1A3A5C")
    H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    BRD    = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    ALT    = PatternFill("solid", fgColor="EEF2FF")
    GRN    = Font(color="1E8449", bold=True, name="Arial", size=10)
    RED    = Font(color="C0392B", bold=True, name="Arial", size=10)

    hdrs = ["Sample","Model","E_corr (V)","i_corr (A/cm²)",
            "βa (mV/dec)","βc (mV/dec)","B (V)",
            "CR (mm/yr)","i_pass (A/cm²)","E_pass (V)",
            "E_trans (V)","R²","RMSE (log)","Status"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = H_FILL; cell.font = H_FONT; cell.border = BRD
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, res in enumerate(results_list, 2):
        p  = res["params"]; ct = res["ct"]
        ok = res.get("r2", 0) > 0.95
        fill = ALT if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        B_val = (p[2]*p[3]) / (2.303*(p[2]+p[3])) if p[2]>0 and p[3]>0 else 0
        ew, rho = res.get("material", (27.92, 7.87))
        icorr_display = res.get("icorr_disp", p[1])
        CR = icorr_display * 3.27 * ew / rho
        vals = [
            res.get("name","?"), CT.name(ct),
            round(res.get("ecorr_disp", p[0]), 5),
            f"{icorr_display:.4e}",
            round(p[2]*1000, 2), round(p[3]*1000, 2), round(B_val, 5), round(CR, 5),
            f"{p[6]:.3e}" if ct in CT.PASS else "—",
            round(p[4], 5) if ct in CT.PASS else "—",
            round(p[7], 5) if ct in CT.TRANS else "—",
            round(res.get("r2", 0), 6), round(res.get("rmse", 0), 6),
            "Good" if ok else "Check",
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=c, value=val)
            cell.fill = fill; cell.border = BRD; cell.alignment = Alignment(horizontal="center")
            if c == 14: cell.font = GRN if ok else RED

    for col in ws.columns:
        w = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+3, 22)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def export_pdf(results_list, png_list):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             rightMargin=2*cm, leftMargin=2*cm,
                             topMargin=2.5*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    ts = ParagraphStyle("T", parent=styles["Title"], fontSize=20,
                         textColor=rl_colors.HexColor("#1A3A5C"), spaceAfter=4)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=12,
                          textColor=rl_colors.HexColor("#2e86de"),
                          spaceBefore=10, spaceAfter=3)
    bs = ParagraphStyle("B", parent=styles["Normal"], fontSize=9, leading=14)

    tbl_s = TableStyle([
        ("BACKGROUND", (0,0),(-1,0), rl_colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",  (0,0),(-1,0), rl_colors.white),
        ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),
         [rl_colors.HexColor("#EEF2FF"), rl_colors.white]),
        ("GRID",       (0,0),(-1,-1), 0.4, rl_colors.HexColor("#BBBBBB")),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("ALIGN",      (2,1),(-1,-1), "RIGHT"),
        ("TOPPADDING", (0,0),(-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1), 3),
    ])

    story = []
    story.append(Paragraph("Polarisation Curve Analysis Report", ts))
    story.append(HRFlowable(width="100%", thickness=2,
                             color=rl_colors.HexColor("#2e86de"), spaceAfter=4))
    story.append(Paragraph(
        f"Polarization Curve Fitter  |  "
        f"{pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}", bs))
    story.append(Spacer(1, 0.5*cm))

    for idx, (res, png) in enumerate(zip(results_list, png_list)):
        p  = res["params"]; ct = res["ct"]
        nm = res.get("name", f"Sample {idx+1}")
        B_val = (p[2]*p[3])/(2.303*(p[2]+p[3])) if p[2]>0 and p[3]>0 else 0
        ew, rho = res.get("material", (27.92, 7.87))
        icorr_display = res.get("icorr_disp", p[1])
        CR = icorr_display * 3.27 * ew / rho

        story.append(Paragraph(f"Sample {idx+1}: {nm}", h2))
        rows = [["Parameter","Symbol","Value","Unit"],
                ["Corrosion potential","E_corr",f"{res.get('ecorr_disp', p[0]):.5f}","V"],
                ["Corrosion current density","i_corr",f"{icorr_display:.4e}","A cm-2"],
                ["Anodic Tafel slope","ba",f"{p[2]*1000:.2f}","mV dec-1"],
                ["Cathodic Tafel slope","bc",f"{p[3]*1000:.2f}","mV dec-1"],
                ["Stern-Geary constant","B",f"{B_val:.5f}","V"],
                ["Corrosion rate","CR",f"{CR:.5f}","mm yr-1"],
                ]
        if ct in CT.PASS:
            rows += [["Passive current density","i_pass",f"{p[6]:.4e}","A cm-2"],
                     ["Passivation potential","E_pass",f"{p[4]:.5f}","V"]]
        if ct in CT.TRANS:
            rows.append(["Transpassive potential","E_trans",f"{p[7]:.5f}","V"])
        rows += [["R² (log-domain)","R2",f"{res.get('r2',0):.6f}","—"],
                 ["RMSE (log-domain)","RMSE",f"{res.get('rmse',0):.6f}","log-units"],
                 ["Model","—",CT.name(ct),"—"],
                 ["Fit status","—","Converged" if res.get("success") else "Check","—"]]
        tbl = Table(rows, colWidths=[6*cm, 2.2*cm, 3.2*cm, 2.6*cm])
        tbl.setStyle(tbl_s)
        story.append(KeepTogether([tbl, Spacer(1, 0.3*cm)]))
        if png:
            story.append(RLImage(io.BytesIO(png), width=15.5*cm, height=11.0*cm))
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width="100%", thickness=0.5,
                                 color=rl_colors.HexColor("#CCCCCC"), spaceAfter=4))

    doc.build(story)
    buf.seek(0); return buf

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for _k in ("results", "figures"):
    if _k not in st.session_state:
        st.session_state[_k] = []

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    st.divider()
    st.markdown("**Material**")
    material = st.selectbox("Material (for CR calculation)",
                            list(MATERIALS.keys()), index=0)
    ew_mat, rho_mat = MATERIALS[material]

    st.markdown("**Data Import**")
    skip_rows  = st.number_input("Skip header rows", 0, 30, 0)
    delimiter  = st.selectbox("CSV delimiter", ["auto",",",";","\t"," "])
    e_col_name = st.text_input("E column (blank = auto)", "")
    i_col_name = st.text_input("i column (blank = auto)", "")
    i_unit     = st.selectbox("Current unit in file",
                              ["A/cm²","mA/cm²","µA/cm²","A/m²"])
    unit_fac   = {"A/cm²":1.0,"mA/cm²":1e-3,"µA/cm²":1e-6,"A/m²":1e-4}[i_unit]
    area       = st.number_input("Electrode area (cm²)", 0.001, 10000.0, 1.0, format="%.4f")

    st.markdown("**Plotting**")
    extend_tafel = st.toggle("Extend Tafel dashed lines to Ecorr", True)
    use_tafel_icorr = st.toggle("Use Tafel intersection for i_corr", True)
    show_regs  = st.toggle("Shade regions", True)
    smooth_pre = st.toggle("Pre-smooth (Savitzky-Golay)", False)
    pub_dpi    = st.slider("Export DPI", 150, 600, 300, 50)

    with st.expander("Advanced (Tafel window detection)"):
        CFG["anod_guard"]   = st.number_input("Anodic guard from Ecorr (V)", 0.0, 0.100, CFG["anod_guard"], 0.001, format="%.3f")
        CFG["cath_guard"]   = st.number_input("Cathodic guard from Ecorr (V)", 0.0, 0.150, CFG["cath_guard"], 0.001, format="%.3f")
        CFG["curvature_max"]= st.number_input("Curvature max |d²log|i|/dE²|", 10.0, 200.0, CFG["curvature_max"], 1.0)
        CFG["lin_frac"]     = st.slider("Linearity fraction (derivative consistency)", 0.4, 0.95, CFG["lin_frac"], 0.05)
        CFG["min_w_ano"]    = st.number_input("Min window points (anodic)", 3, 20, CFG["min_w_ano"])
        CFG["min_w_cat"]    = st.number_input("Min window points (cathodic)", 4, 25, CFG["min_w_cat"])
        CFG["beta_min"]     = st.number_input("β min (V/dec)", 0.005, 0.100, CFG["beta_min"], 0.005, format="%.3f")
        CFG["beta_max_a"]   = st.number_input("βa max (V/dec)", 0.05, 0.40, CFG["beta_max_a"], 0.005)
        CFG["beta_max_c"]   = st.number_input("βc max (V/dec)", 0.05, 0.40, CFG["beta_max_c"], 0.005)

    st.markdown("**Fitting**")
    force_ct_val = st.selectbox("Force model (auto = best AICc)",
                                ["auto","A","AD","P","PT","F"],
                                key="force_ct_choice")

    st.divider()
    if st.button("🗑 Clear all", use_container_width=True):
        st.session_state.results = []
        st.session_state.figures = []
        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-header">⚡ Polarization Curve Fitter</div>',
            unsafe_allow_html=True)

tab_fit, tab_res, tab_cmp, tab_help = st.tabs(
    ["📂 Upload & Fit", "📊 Results & Export", "📋 Compare", "ℹ️ Help"])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — UPLOAD & FIT
# ─────────────────────────────────────────────────────────────────────────────
with tab_fit:
    c1, c2 = st.columns([1.2, 0.8])
    with c1:
        st.markdown("### 📁 Upload Data")
        uploaded_files = st.file_uploader(
            "CSV / TXT / XLSX  (signed current: cathodic < 0, anodic > 0)",
            type=["csv","txt","xlsx","xls"],
            accept_multiple_files=True)
    with c2:
        st.markdown("### 🏷️ Sample")
        sample_name = st.text_input("Sample label", "Sample 1")

    if uploaded_files:
        for idx, uf in enumerate(uploaded_files):
            st.markdown(f"---\n#### 📄 `{uf.name}`")
            with st.container():

                # Load
                try:
                    df_raw = load_file(uf)
                except Exception as ex:
                    st.error(f"Load error: {ex}"); continue

                # Column selection
                try:
                    ec_auto, ic_auto, _ = _auto_cols(df_raw)
                    auto_ok = True
                except:
                    ec_auto = ic_auto = None; auto_ok = False

                num_cols = [c for c in df_raw.columns
                            if pd.api.types.is_numeric_dtype(df_raw[c])]
                cc1, cc2 = st.columns(2)
                with cc1:
                    e_sel = st.selectbox(f"E column [{uf.name}]", num_cols,
                        index=num_cols.index(ec_auto) if auto_ok and ec_auto in num_cols else 0,
                        key=f"ec_{idx}")
                with cc2:
                    i_sel = st.selectbox(f"i column [{uf.name}]", num_cols,
                        index=num_cols.index(ic_auto) if auto_ok and ic_auto in num_cols else min(1,len(num_cols)-1),
                        key=f"ic_{idx}")

                # Build arrays
                E_raw = df_raw[e_sel].values.astype(float)
                i_raw = df_raw[i_sel].values.astype(float)
                i_obs = i_raw * unit_fac / area
                ok_mask = np.isfinite(E_raw) & np.isfinite(i_obs)
                E_raw, i_obs = E_raw[ok_mask], i_obs[ok_mask]
                srt = np.argsort(E_raw); E, i = E_raw[srt], i_obs[srt]

                # Preview
                p1, p2 = st.columns([1, 1.6])
                sc = np.where(np.diff(np.sign(i)))[0]
                with p1:
                    st.markdown(f"**{len(E)} pts** | "
                                f"E: [{E.min():.4f}, {E.max():.4f}] V")
                    if len(sc) > 0:
                        Ec_pre = E[sc[0]] - i[sc[0]]*(E[sc[0]+1]-E[sc[0]])/(i[sc[0]+1]-i[sc[0]])
                        st.success(f"✓ Zero-crossing at E ≈ {Ec_pre:.4f} V  "
                                   f"({np.sum(i<0)} cat / {np.sum(i>0)} ano)")
                    else:
                        st.warning("No sign change — verify sign convention")
                    st.dataframe(df_raw.head(5), use_container_width=True, height=160)

                with p2:
                    with plt.rc_context(PLT_RC):
                        fp, ap = plt.subplots(figsize=(6, 3.8))
                        ap.scatter(E, slog(i), s=7, color="#5a7fa8", alpha=0.65, rasterized=True)
                        ap.set_xlabel("E (V)"); ap.set_ylabel("log|i|")
                        ap.set_title("Raw Data Preview", fontsize=10)
                        ap.grid(True, ls="--", alpha=0.4)
                        if len(sc) > 0:
                            ap.axvline(Ec_pre, color="#e84393", ls="--", lw=1,
                                       label=f"E_corr≈{Ec_pre:.4f}V")
                            ap.legend(fontsize=8)
                        fp.tight_layout()
                    st.pyplot(fp, use_container_width=True)
                    plt.close(fp)

                # ── FIT BUTTON ──────────────────────────────────────────────
                if st.button(f"🚀 Run Full Pipeline · {uf.name}",
                             key=f"btn_{idx}", type="primary",
                             use_container_width=True):
                    import time; t0 = time.time()

                    if smooth_pre:
                        w = min(9, len(i)//2*2-1)
                        if w >= 5:
                            i = savgol_filter(i, w, 3, mode="interp")

                    prog     = st.progress(0, text="Stage 1 — Detecting E_corr…")
                    log_area = st.empty()
                    log_lines = []
                    def log(msg):
                        log_lines.append(msg)
                        log_area.markdown("  \n".join(log_lines))

                    # Stage 1 — E_corr
                    Ecorr, _ = detect_ecorr(E, i)
                    log(f"✅ **Stage 1** — E_corr = `{Ecorr:.5f}` V")
                    prog.progress(15, text="Stage 2 — Cathodic branch fit…")

                    # Stage 2 — Cathodic
                    cat_res = fit_cathodic(E, i, Ecorr)
                    win_c_str = f"  win=[`{cat_res['win_c'][0]:.4f}`, `{cat_res['win_c'][1]:.4f}`]V" if 'win_c' in cat_res else ""
                    log(f"✅ **Stage 2** — βc(local) = `{cat_res['bc']*1000:.0f}` mV/dec  "
                        f"i_corr ≈ `{cat_res['icorr']:.2e}`  "
                        f"R²_cat = `{cat_res['r2']:.4f}`  "
                        f"diff_limit = `{'yes' if cat_res['has_diff'] else 'no'}`"
                        f"{win_c_str}")
                    prog.progress(30, text="Stage 3 — Anodic branch fit…")

                    # Stage 3 — Anodic
                    an_res = fit_anodic(E, i, Ecorr)
                    epeak_str = f"  Epeak=`{an_res['Epeak']:.4f}V`" if an_res.get('Epeak') else ""
                    epass_str = f"  E_pass=`{an_res['Epass']:.4f}V`  ip=`{an_res['ip']:.2e}`" if an_res['has_passive'] else ""
                    etrans_str = f"  E_trans=`{an_res['Etrans']:.4f}V`" if an_res.get('has_trans') and an_res.get('Etrans') else ""
                    log(f"✅ **Stage 3** — βa(local) = `{an_res['ba']*1000:.0f}` mV/dec"
                        f"{epeak_str}"
                        f"  passive = `{'yes' if an_res['has_passive'] else 'no'}`{epass_str}"
                        f"  transpassive = `{'yes' if an_res['has_trans'] else 'no'}`{etrans_str}"
                        f"  R²_ano = `{an_res['r2']:.4f}`")
                    prog.progress(45, text="Stage 4 — Classifying curve type…")

                    # Stage 4 — Classify
                    ct_detected = classify_curve(cat_res, an_res)
                    log(f"🔍 **Stage 4** — Detected: **{CT.name(ct_detected)}**  "
                        f"({CT.nfree(ct_detected)} free parameters)")
                    prog.progress(50, text="Stage 5 — Global optimisation…")

                    # Stage 5 — Global optimisation
                    E_lo = float(E.min()); E_hi = float(E.max()); E_sp = E_hi - E_lo

                    force_choice = st.session_state.get("force_ct_choice", "auto")
                    if force_choice == "auto":
                        # Build candidate set. F (Full: passive+diffusion) is always
                        # tried when both passive and diffusion features are present — the
                        # local detection is conservative and can miss one of them.
                        candidates = []
                        if an_res["has_passive"] and cat_res["has_diff"]:
                            # Both features: F is the most likely correct model
                            candidates = [CT.F, CT.PT, CT.P, CT.AD]
                        elif an_res["has_passive"]:
                            candidates = [CT.PT, CT.P, CT.F]
                        elif cat_res["has_diff"]:
                            candidates = [CT.AD, CT.A, CT.P]
                        else:
                            candidates = [ct_detected, CT.P]
                        # Always include the auto-detected type
                        if ct_detected not in candidates:
                            candidates.insert(0, ct_detected)
                        seen = set(); uniq = []
                        for c in candidates:
                            if c not in seen:
                                uniq.append(c); seen.add(c)
                        candidates = uniq
                    else:
                        candidates = [force_choice]

                    all_res = []
                    n_cand  = len(candidates)
                    for k_c, ct_try in enumerate(candidates):
                        prog.progress(50 + int(40 * k_c / max(n_cand,1)),
                                      text=f"Optimising: {CT.name(ct_try)}…")
                        p0 = _make_p0(Ecorr, cat_res, an_res, ct_try, E_hi)
                        lo, hi = _build_bounds(Ecorr, cat_res, an_res,
                                               ct_try, E_lo, E_hi, E_sp)
                        p0 = np.clip(p0, lo, hi)
                        bp, r2v, aic_v, sse = global_polish(E, i, p0, ct_try, lo, hi)
                        all_res.append(dict(ct=ct_try, r2=r2v, aicc=aic_v,
                                            params=bp, success=r2v > 0.90))
                        log(f"  · {CT.name(ct_try):35s} R² = `{r2v:.6f}`  AICc = `{aic_v:.1f}`")

                    # AICc selection — lowest AICc wins, but allow a simpler model
                    # only if it costs < 0.5% R² (not 0.2% — too tight for noisy data).
                    all_res.sort(key=lambda x: x["aicc"])
                    best_r = all_res[0]
                    for r in all_res:
                        if (CT.nfree(r["ct"]) < CT.nfree(best_r["ct"])
                                and best_r["r2"] - r["r2"] < 0.005):
                            best_r = r; break

                    best_p  = best_r["params"]
                    best_ct = best_r["ct"]
                    r2_fin  = best_r["r2"]

                    log(f"🏆 **Stage 5 complete** — Best model: **{CT.name(best_ct)}**  "
                        f"R² = `{r2_fin:.6f}`  "
                        f"(elapsed `{time.time()-t0:.1f}` s)")
                    prog.progress(95, text="Building figure…")

                    # Tafel intersection
                    taf = tafel_intersection(cat_res, an_res)

                    # ── Figure ──────────────────────────────────────────────
                    try:
                        fig, r2_fig, rmse_fig, icorr_disp, ecorr_disp = make_figure(
                            E, i, best_p, best_ct, sample_name or uf.name,
                            cat_res, an_res, Ecorr, taf=taf,
                            extend_tafel=extend_tafel, use_tafel_icorr=use_tafel_icorr,
                            show_regions=show_regs, dpi=pub_dpi)
                        fig.tight_layout(rect=[0, 0, 1, 0.97])

                        buf_png = io.BytesIO()
                        fig.savefig(buf_png, dpi=pub_dpi, bbox_inches="tight",
                                    facecolor="white"); buf_png.seek(0)
                        png_bytes = buf_png.read()

                        buf_svg = io.BytesIO()
                        fig.savefig(buf_svg, format="svg", bbox_inches="tight",
                                    facecolor="white"); buf_svg.seek(0)
                        svg_bytes = buf_svg.read()

                        res_rec = dict(
                            name=sample_name or uf.name,
                            params=best_p, ct=best_ct,
                            r2=r2_fin, rmse=rmse_fig,
                            success=r2_fin > 0.90,
                            material=(ew_mat, rho_mat),
                            all_candidates=all_res,
                            icorr_disp=icorr_disp, ecorr_disp=ecorr_disp
                        )
                        st.session_state.results.append(res_rec)
                        st.session_state.figures.append({
                            "png": png_bytes, "svg": svg_bytes,
                            "name": res_rec["name"]})

                        st.pyplot(fig, use_container_width=True)
                        plt.close(fig)
                    except Exception as ex:
                        st.error(f"Figure error: {ex}")
                        st.code(traceback.format_exc())

                    # ── Metrics ──────────────────────────────────────────────
                    p = best_p
                    B_val = (p[2]*p[3])/(2.303*(p[2]+p[3])) if p[2]>0 and p[3]>0 else 0
                    CR    = icorr_disp * 3.27 * ew_mat / rho_mat

                    # Runaway detection (from busbar_tafel_fit.py ba_valid check)
                    ba_valid = float(p[2]) < 0.240   # > 240 mV/dec = unphysical (was 0.495)
                    bc_valid = float(p[3]) < 0.270   # > 270 mV/dec = unphysical (was 0.495)
                    if not ba_valid or not bc_valid:
                        st.warning(
                            f"⚠️ {'βa' if not ba_valid else ''}"
                            f"{' and ' if not ba_valid and not bc_valid else ''}"
                            f"{'βc' if not bc_valid else ''} hit the 500 mV/dec upper bound — "
                            "no reliable linear Tafel region was found for that branch. "
                            "The corrosion current is estimated from the other branch only. "
                            "Consider adjusting the guard or window in Advanced settings.")

                    st.markdown("#### 📐 Fitted Parameters")
                    mc = st.columns(5)
                    mc[0].metric("E_corr (V)",       f"{ecorr_disp:.5f}")
                    mc[1].metric("i_corr (A/cm²)",   f"{icorr_disp:.4e}")
                    mc[2].metric("βa (mV/dec)",       f"{p[2]*1000:.1f}" + ("" if ba_valid else " ⚠️"))
                    mc[3].metric("βc (mV/dec)",       f"{p[3]*1000:.1f}" + ("" if bc_valid else " ⚠️"))
                    mc[4].metric("B (V)",             f"{B_val:.5f}")

                    mc2 = st.columns(5)
                    mc2[0].metric("CR (mm/yr)",      f"{CR:.5f}")
                    mc2[1].metric("R²",              f"{r2_fin:.5f}",
                                  "Excellent" if r2_fin>0.99 else
                                  "Good" if r2_fin>0.95 else "⚠ Check")
                    if best_ct in CT.PASS:
                        mc2[2].metric("i_pass (A/cm²)", f"{p[6]:.4e}")
                        mc2[3].metric("E_pass (V)",     f"{p[4]:.5f}")
                    if best_ct in CT.TRANS:
                        mc2[4].metric("E_trans (V)",    f"{p[7]:.5f}")

                    # ── Per-region local linear fit summary ──────────────────
                    st.markdown("#### 🔬 Local Linear Fits by Region")
                    rr_cols = st.columns(3)
                    with rr_cols[0]:
                        st.markdown("**Cathodic Tafel**")
                        if "win_c" in cat_res:
                            st.markdown(
                                f"- Window: `{cat_res['win_c'][0]:.4f}` → `{cat_res['win_c'][1]:.4f}` V  \n"
                                f"- βc = **{cat_res['bc']*1000:.1f} mV/dec**  \n"
                                f"- R² = `{cat_res['r2']:.4f}`  \n"
                                f"- Diff. limit: `{'yes' if cat_res['has_diff'] else 'no'}`"
                                + (f"  \n- iL ≈ `{cat_res['iL']:.3e}` A/cm²" if cat_res['has_diff'] else ""))
                    with rr_cols[1]:
                        st.markdown("**Anodic Active Tafel**")
                        if "win_a" in an_res:
                            active_up2 = an_res.get("Epeak") or (float(p[4]) if best_ct in CT.PASS else None)
                            wa1_disp = min(float(an_res["win_a"][1]), active_up2) if active_up2 else float(an_res["win_a"][1])
                            st.markdown(
                                f"- Window: `{an_res['win_a'][0]:.4f}` → `{wa1_disp:.4f}` V  \n"
                                f"- βa = **{an_res['ba']*1000:.1f} mV/dec**  \n"
                                f"- R² = `{an_res['r2']:.4f}`  \n"
                                + (f"- Epeak ≈ `{an_res['Epeak']:.4f}` V" if an_res.get('Epeak') else ""))
                    with rr_cols[2]:
                        if best_ct in CT.PASS:
                            st.markdown("**Passive Region**")
                            st.markdown(
                                f"- E_pass = `{p[4]:.4f}` V (model)  \n"
                                f"- i_pass = `{p[6]:.3e}` A/cm²  \n"
                                + (f"- E_trans = `{p[7]:.4f}` V" if best_ct in CT.TRANS else "- Transpassive: not detected"))
                        else:
                            st.markdown("**No passive region detected**")

                    if len(all_res) > 1:
                        st.markdown("**🏆 Model Comparison (AICc)**")
                        cmp_rows = [{
                            "Model": CT.name(r["ct"]),
                            "Free params": CT.nfree(r["ct"]),
                            "R²": f"{r['r2']:.6f}",
                            "AICc": f"{r['aicc']:.1f}",
                            "Selected": "✅" if r["ct"]==best_ct else ""
                        } for r in sorted(all_res, key=lambda x: x["aicc"])]
                        st.dataframe(pd.DataFrame(cmp_rows),
                                     use_container_width=True, hide_index=True)

                    d1, d2 = st.columns(2)
                    d1.download_button("⬇ PNG", png_bytes,
                                       f"{sample_name}.png", "image/png")
                    d2.download_button("⬇ SVG", svg_bytes,
                                       f"{sample_name}.svg", "image/svg+xml")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — RESULTS & EXPORT
# ─────────────────────────────────────────────────────────────────────────────
with tab_res:
    if not st.session_state.results:
        st.info("No results yet.")
    else:
        rows = []
        for r in st.session_state.results:
            p = r["params"]; ct = r["ct"]
            B  = (p[2]*p[3])/(2.303*(p[2]+p[3])) if p[2]>0 and p[3]>0 else 0
            ew, rho = r.get("material",(27.92,7.87))
            icorr_disp = r.get("icorr_disp", p[1])
            CR = icorr_disp * 3.27 * ew / rho
            rows.append({"Sample":r.get("name","?"),
                          "Model":CT.name(ct),
                          "E_corr (V)":f"{r.get('ecorr_disp', p[0]):.5f}",
                          "i_corr (A/cm²)":f"{icorr_disp:.4e}",
                          "βa (mV/dec)":f"{p[2]*1000:.1f}",
                          "βc (mV/dec)":f"{p[3]*1000:.1f}",
                          "B (V)":f"{B:.5f}",
                          "CR (mm/yr)":f"{CR:.5f}",
                          "i_pass":f"{p[6]:.3e}" if ct in CT.PASS else "—",
                          "E_pass (V)":f"{p[4]:.5f}" if ct in CT.PASS else "—",
                          "E_trans (V)":f"{p[7]:.5f}" if ct in CT.TRANS else "—",
                          "R²":f"{r.get('r2',0):.5f}",
                          "RMSE":f"{r.get('rmse',0):.5f}",
                          "Status":"✓ Good" if r.get("r2",0)>0.95 else "⚠ Check"})
        st.dataframe(pd.DataFrame(rows), use_container_width=True)
        st.divider()

        ec1, ec2, ec3 = st.columns(3)
        ec1.download_button("📥 Excel (.xlsx)",
            data=export_excel(st.session_state.results),
            file_name="polarization_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
        ec2.download_button("📥 PDF Report",
            data=export_pdf(st.session_state.results,
                            [f["png"] for f in st.session_state.figures]),
            file_name="polarization_report.pdf", mime="application/pdf",
            use_container_width=True)
        zb = io.BytesIO()
        with zipfile.ZipFile(zb, "w") as zf:
            for fd in st.session_state.figures:
                zf.writestr(f"{fd['name']}.png", fd["png"])
                zf.writestr(f"{fd['name']}.svg", fd["svg"])
        zb.seek(0)
        ec3.download_button("📥 Figures (.zip)", data=zb,
            file_name="polarization_figures.zip", mime="application/zip",
            use_container_width=True)

        st.markdown("### Fitted Figures")
        for fd in st.session_state.figures:
            st.markdown(f"**{fd['name']}**")
            st.image(fd["png"], use_container_width=True)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — COMPARISON
# ─────────────────────────────────────────────────────────────────────────────
with tab_cmp:
    if len(st.session_state.results) < 2:
        st.info("Fit ≥ 2 samples to enable comparison.")
    else:
        with plt.rc_context(PLT_RC):
            fig_c, axes = plt.subplots(1, 3, figsize=(15, 5), dpi=120)
            names = [r.get("name","?") for r in st.session_state.results]
            for idx, res in enumerate(st.session_state.results):
                col = PALETTE[idx % len(PALETTE)]
                p   = res["params"]; ct = res["ct"]
                lbl = res.get("name", f"S{idx+1}")
                E_pl = np.linspace(min(p[0]-1.5, -1.5), max(p[0]+1.5, 1.5), 3000)
                try:
                    i_pl = pol_model(E_pl, p, ct)
                    axes[0].plot(E_pl, slog(i_pl), color=col, lw=2, label=lbl)
                    axes[0].axvline(res.get("ecorr_disp", p[0]), color=col, ls=":", lw=0.9, alpha=0.6)
                except:
                    pass
                axes[1].bar(idx, res.get("icorr_disp", p[1]), color=col, alpha=0.85)
                axes[2].bar(idx-0.2, p[2]*1000, 0.38, color=col, alpha=0.85)
                axes[2].bar(idx+0.2, p[3]*1000, 0.38, color=col, alpha=0.45, hatch="//")

            axes[0].set_xlabel("E (V)"); axes[0].set_ylabel("log|i| (A/cm²)")
            axes[0].set_title("Evans Diagram Overlay", fontweight="bold")
            axes[0].legend(fontsize=8); axes[0].grid(True, ls="--", alpha=0.35)
            axes[0].set_facecolor("#fafbff")

            for ax, ttl, yl in zip(axes[1:],
                ["i_corr Comparison", "Tafel Slopes (filled=βa, hatch=βc)"],
                ["i_corr (A/cm²)", "Tafel slope (mV/dec)"]):
                ax.set_xticks(range(len(names)))
                ax.set_xticklabels(names, rotation=18, ha="right")
                ax.set_ylabel(yl); ax.set_title(ttl, fontweight="bold")
                ax.grid(True, axis="y", ls="--", alpha=0.35)
                ax.set_facecolor("#fafbff")
            axes[1].set_yscale("log")
            fig_c.tight_layout()
        st.pyplot(fig_c, use_container_width=True)
        plt.close(fig_c)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — HELP
# ─────────────────────────────────────────────────────────────────────────────
with tab_help:
    st.markdown("""
### Algorithm Fixes in v2

| Fix | Details |
|---|---|
| **Cathodic proximity score** | Gaussian centred at 150 mV from E_corr (σ=120 mV). Previously rewarded windows *near* E_corr — the mixed-control zone — instead of the true Tafel region 50–300 mV away. |
| **Cathodic scoring** | R²² weighting and direct curvature penalty added to sharpen window discrimination. |
| **Cathodic diff_ok guard** | Threshold lowered from 0.35→0.20 × slope_mag. The original threshold over-rejected valid cathodic windows with natural noise. |
| **Cathodic i_corr estimate** | Uses `max(tafel_extrap, near_Ecorr × 0.5)` instead of `min(tafel_extrap, near_Ecorr × 10)`. The old cap systematically under-estimated i_corr when the Tafel line was correctly picked far from E_corr. |
| **Anodic proximity tau** | Widened from 0.12 V → 0.20 V so active dissolution regions spanning 100–200 mV (stainless, Ni, Ti) are not penalised. |
| **Anodic fallback ok mask** | Retains `beta_ok` in fallback to prevent selection of windows with unphysical Tafel slopes. |
| **Linear-scale panel ylim** | Now based on |i| within ±150 mV of E_corr, not the global 95th percentile. Passive/transpassive current peaks no longer compress the Ecorr region. |

### Why the cathodic dashed line can look shorter
- Diffusion-limited plateaus flatten the cathodic branch; only a region 50–300 mV
  from E_corr is truly linear in the Tafel sense.
- Curvature near E_corr (IR/mixed control) is excluded by the curvature guard.
- Enable "Extend Tafel dashed lines" in the sidebar to project the fitted window
  faintly to E_corr for visual reference.

### Troubleshooting

| Symptom | Try |
|---|---|
| Cathodic line too short / wrong slope | Reduce *Cathodic guard* to 0.010 V; raise *Curvature max* to 60 |
| Anodic line not detected | Reduce *Anodic guard* to 0.005 V; lower *Linearity fraction* to 0.60 |
| i_corr looks wrong | Toggle *Use Tafel intersection for i_corr*; check Tafel intersection marker |
| Very noisy data | Enable *Pre-smooth* in sidebar |

### Fitting Pipeline

| Stage | What happens |
|---|---|
| **1 — E_corr detection** | Interpolated zero-crossing of signed current |
| **2 — Cathodic fit** | Vectorized sliding regression; Gaussian proximity scoring at 150 mV from E_corr; Theil–Sen/Huber refinement |
| **3 — Anodic fit** | Same for anodic branch; wider proximity window (tau=0.20 V); passive plateau detection |
| **4 — Classification** | Curve type inferred from detected features |
| **5 — Global polish** | Physics-informed p₀ → DE → L-BFGS-B → Powell (if needed) |
| **6 — AICc selection** | Multiple candidate models compared; parsimony-penalised |
""")
